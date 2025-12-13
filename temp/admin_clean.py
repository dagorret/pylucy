from django.contrib import admin
from django.http import HttpResponse
from .models import (
    ImportLog, Course, Category, Enrol, UserEnrolment, 
    User, Group, GroupMember, UserLastAccess, RoleAssignment, Context
)

# Action para exportar a Excel
def export_to_excel(modeladmin, request, queryset):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = modeladmin.model._meta.verbose_name_plural[:31]
    
    fields = [f for f in modeladmin.model._meta.fields if f.name not in ['id']]
    headers = [f.verbose_name for f in fields]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for obj in queryset:
        row = [getattr(obj, f.name) for f in fields]
        ws.append(row)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{modeladmin.model._meta.verbose_name_plural}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

export_to_excel.short_description = "📊 Exportar a Excel"


@admin.register(ImportLog)
class ImportLogAdmin(admin.ModelAdmin):
    list_display = ['table_name', 'status', 'records_imported', 'started_at']
    list_filter = ['status', 'table_name']
    readonly_fields = ['table_name', 'started_at', 'completed_at', 'records_imported', 'status', 'error_message']


@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'shortname', 'fullname', 'category', 'visible']
    list_filter = ['visible', 'category']
    search_fields = ['shortname', 'fullname']
    actions = [export_to_excel]


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'name', 'parent', 'visible']
    list_filter = ['visible']
    search_fields = ['name']
    actions = [export_to_excel]


@admin.register(Enrol)
class EnrolAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'courseid', 'enrol', 'status']
    list_filter = ['enrol', 'status']
    actions = [export_to_excel]


@admin.register(UserEnrolment)
class UserEnrolmentAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'enrolid', 'status']
    list_filter = ['status']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'username', 'firstname', 'lastname', 'email']
    search_fields = ['username', 'firstname', 'lastname', 'email']
    list_filter = ['suspended']
    actions = [export_to_excel]


@admin.register(Group)
class GroupAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'name', 'courseid']
    search_fields = ['name']
    actions = [export_to_excel]


@admin.register(GroupMember)
class GroupMemberAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'groupid', 'userid']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(UserLastAccess)
class UserLastAccessAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'courseid', 'timeaccess']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(RoleAssignment)
class RoleAssignmentAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'roleid', 'contextid']
    list_filter = ['roleid']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(Context)
class ContextAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'contextlevel', 'instanceid', 'depth']
    list_filter = ['contextlevel']
    list_per_page = 100
    actions = [export_to_excel]
