"""
Configuración del Admin de Django para Moodle Data
"""
from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path, reverse
from django.utils.html import format_html
from django.contrib import messages
from django.core.management import call_command
from django.http import HttpResponse
from django.template.response import TemplateResponse
from .models import (
    ImportLog, Course, Category, Enrol, UserEnrolment, 
    User, Group, GroupMember, UserLastAccess, RoleAssignment, Context
)
import io


# Personalizar el AdminSite para agregar botones en el index
class MoodleAdminSite(admin.AdminSite):
    site_header = "Moodle Stats - Administración"
    site_title = "Moodle Stats Admin"
    index_title = "Panel de Control de Datos de Moodle"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-all/', self.admin_view(self.export_all_view), name='export_all'),
            path('import-all/', self.admin_view(self.import_all_view), name='import_all'),
        ]
        return custom_urls + urls
    
    def index(self, request, extra_context=None):
        """Sobrescribir index para agregar botones personalizados"""
        extra_context = extra_context or {}
        extra_context['custom_actions'] = True
        return super().index(request, extra_context)
    
    def export_all_view(self, request):
        """Exporta todas las tablas desde Moodle"""
        if request.method == 'POST':
            try:
                out = io.StringIO()
                call_command('export_moodle', '--tables', 'all', stdout=out)
                
                messages.success(request, '✓ Exportación completada exitosamente')
                output = out.getvalue()
                for line in output.split('\n'):
                    if line.strip():
                        messages.info(request, line)
                
            except Exception as e:
                messages.error(request, f'Error en la exportación: {str(e)}')
            
            return redirect('admin:index')
        
        # GET - Mostrar formulario
        context = {
            'title': 'Exportar todas las tablas desde Moodle',
            'site_header': self.site_header,
            'site_title': self.site_title,
            'has_permission': True,
            'tables': [
                'courses', 'categories', 'enrol', 'user_enrolments', 'users',
                'groups', 'groups_members', 'user_lastaccess', 'role_assignments', 'context'
            ]
        }
        return TemplateResponse(request, 'admin/export_all_confirm.html', context)
    
    def import_all_view(self, request):
        """Importa todas las tablas a Django"""
        if request.method == 'POST':
            clear = request.POST.get('clear') == 'on'
            
            try:
                out = io.StringIO()
                if clear:
                    call_command('import_moodle', '--tables', 'all', '--clear', stdout=out)
                else:
                    call_command('import_moodle', '--tables', 'all', stdout=out)
                
                messages.success(request, '✓ Importación completada exitosamente')
                output = out.getvalue()
                for line in output.split('\n'):
                    if line.strip():
                        messages.info(request, line)
                
            except Exception as e:
                messages.error(request, f'Error en la importación: {str(e)}')
            
            return redirect('admin:index')
        
        # GET - Mostrar formulario
        context = {
            'title': 'Importar todas las tablas a Django',
            'site_header': self.site_header,
            'site_title': self.site_title,
            'has_permission': True,
            'tables': [
                'courses', 'categories', 'enrol', 'user_enrolments', 'users',
                'groups', 'groups_members', 'user_lastaccess', 'role_assignments', 'context'
            ]
        }
        return TemplateResponse(request, 'admin/import_all_confirm.html', context)


# Usar el AdminSite personalizado
admin_site = MoodleAdminSite(name='moodleadmin')


# Action para exportar a Excel
def export_to_excel(modeladmin, request, queryset):
    """Exporta los registros seleccionados a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = modeladmin.model._meta.verbose_name_plural[:31]
    
    fields = [f for f in modeladmin.model._meta.fields if f.name not in ['id']]
    headers = [f.verbose_name for f in fields]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for obj in queryset:
        row = [getattr(obj, f.name) for f in fields]
        ws.append(row)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{modeladmin.model._meta.verbose_name_plural}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

export_to_excel.short_description = "📊 Exportar seleccionados a Excel"


# Registrar modelos en el admin_site personalizado

@admin_site.register(ImportLog)
class ImportLogAdmin(admin.ModelAdmin):
    list_display = ['id', 'table_name', 'status_badge', 'records_imported', 'started_at', 'duration_display']
    list_filter = ['status', 'table_name', 'started_at']
    search_fields = ['table_name', 'error_message']
    readonly_fields = ['table_name', 'started_at', 'completed_at', 'records_imported', 'status', 'error_message']
    
    def status_badge(self, obj):
        colors = {'running': 'blue', 'completed': 'green', 'failed': 'red'}
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; border-radius: 3px;">{}</span>',
            colors.get(obj.status, 'gray'), obj.get_status_display()
        )
    status_badge.short_description = 'Estado'
    
    def duration_display(self, obj):
        if obj.completed_at and obj.started_at:
            duration = obj.completed_at - obj.started_at
            total_seconds = int(duration.total_seconds())
            minutes, seconds = divmod(total_seconds, 60)
            return f"{minutes}m {seconds}s"
        return "-"
    duration_display.short_description = 'Duración'
    
    def has_add_permission(self, request):
        return False


@admin_site.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'shortname', 'fullname', 'category', 'visible', 'imported_at']
    list_filter = ['visible', 'category', 'imported_at']
    search_fields = ['shortname', 'fullname']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 50
    actions = [export_to_excel]


@admin_site.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'name', 'parent', 'path', 'visible', 'imported_at']
    list_filter = ['visible', 'parent', 'imported_at']
    search_fields = ['name', 'path']
    readonly_fields = ['moodle_id', 'imported_at']
    actions = [export_to_excel]


@admin_site.register(Enrol)
class EnrolAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'courseid', 'enrol', 'status', 'imported_at']
    list_filter = ['enrol', 'status', 'imported_at']
    search_fields = ['courseid']
    readonly_fields = ['moodle_id', 'imported_at']
    actions = [export_to_excel]


@admin_site.register(UserEnrolment)
class UserEnrolmentAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'enrolid', 'status', 'imported_at']
    list_filter = ['status', 'imported_at']
    search_fields = ['userid', 'enrolid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin_site.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'username', 'full_name', 'email', 'suspended', 'imported_at']
    list_filter = ['suspended', 'deleted', 'country', 'imported_at']
    search_fields = ['username', 'firstname', 'lastname', 'email']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 50
    actions = [export_to_excel]
    
    def full_name(self, obj):
        return f"{obj.lastname}, {obj.firstname}"
    full_name.short_description = 'Nombre Completo'


@admin_site.register(Group)
class GroupAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'name', 'courseid', 'idnumber', 'imported_at']
    list_filter = ['courseid', 'imported_at']
    search_fields = ['name', 'idnumber']
    readonly_fields = ['moodle_id', 'imported_at']
    actions = [export_to_excel]


@admin_site.register(GroupMember)
class GroupMemberAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'groupid', 'userid', 'imported_at']
    list_filter = ['imported_at']
    search_fields = ['groupid', 'userid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin_site.register(UserLastAccess)
class UserLastAccessAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'courseid', 'timeaccess', 'imported_at']
    list_filter = ['imported_at']
    search_fields = ['userid', 'courseid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin_site.register(RoleAssignment)
class RoleAssignmentAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'roleid', 'contextid', 'timemodified', 'imported_at']
    list_filter = ['roleid', 'imported_at']
    search_fields = ['userid', 'contextid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin_site.register(Context)
class ContextAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'contextlevel', 'instanceid', 'depth', 'imported_at']
    list_filter = ['contextlevel', 'depth', 'imported_at']
    search_fields = ['instanceid', 'path']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]
