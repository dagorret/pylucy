from django import forms
from django.contrib import admin

from .constants import CARRERAS_CHOICES, MODALIDADES_CHOICES
from .models import (
    Carrera,
    CursoIngreso,
    _normalizar_carreras,
    _normalizar_comisiones,
    _normalizar_modalidades,
)
from .widgets import CarrerasTagWidget


@admin.register(Carrera)
class CarreraAdmin(admin.ModelAdmin):
    """Admin para el modelo Carrera."""
    list_display = ("nombre_completo", "codigo", "id_uti", "activo", "updated_at")
    list_filter = ("activo",)
    search_fields = ("nombre_completo", "codigo")
    ordering = ("nombre_completo",)
    readonly_fields = ("created_at", "updated_at")

    fieldsets = (
        ("Información de la Carrera", {
            "fields": ("nombre_completo", "codigo", "id_uti", "activo")
        }),
        ("Metadatos", {
            "fields": ("created_at", "updated_at"),
            "classes": ("collapse",)
        }),
    )


class BaseCursoAdmin(admin.ModelAdmin):
    list_display = (
        "nombre",
        "curso_moodle",
        "comisiones_display",
        "carreras_display",
        "modalidades_display",
        "activo",
    )
    list_filter = ("activo",)
    search_fields = ("nombre", "curso_moodle", "carreras", "modalidades", "comisiones")
    ordering = ("nombre",)
    readonly_fields = ()

    def carreras_display(self, obj):
        return ", ".join(obj.carreras)

    carreras_display.short_description = "Carreras"

    def modalidades_display(self, obj):
        label_map = dict(MODALIDADES_CHOICES)
        return ", ".join(label_map.get(mod, mod) for mod in obj.modalidades)

    modalidades_display.short_description = "Modalidades"

    def comisiones_display(self, obj):
        return ", ".join(obj.comisiones)

    comisiones_display.short_description = "Comisiones"

    def save_model(self, request, obj, form, change):
        # Normaliza carreras si viene como texto
        obj.carreras = _normalizar_carreras(obj.carreras)
        obj.modalidades = _normalizar_modalidades(obj.modalidades)
        obj.comisiones = _normalizar_comisiones(obj.comisiones)
        super().save_model(request, obj, form, change)


class TagsField(forms.MultipleChoiceField):
    """Permite valores fuera de las choices para crear tags nuevos."""

    def validate(self, value):
        if self.required and not value:
            raise forms.ValidationError(self.error_messages["required"], code="required")
        # No validamos contra las choices para permitir tags nuevos.


class BaseCursoForm(forms.ModelForm):
    carreras = TagsField(
        choices=CARRERAS_CHOICES,
        widget=CarrerasTagWidget(attrs={"data-placeholder": "Escribe o selecciona carreras"}),
    )
    modalidades = TagsField(
        choices=MODALIDADES_CHOICES,
        widget=CarrerasTagWidget(attrs={"data-placeholder": "Selecciona modalidad (PRES/DIST)"}),
    )
    comisiones = TagsField(
        choices=(),
        widget=CarrerasTagWidget(attrs={"data-placeholder": "Escribe comisiones y presiona Enter"}),
    )

    class Meta:
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._bootstrap_tag_field("carreras", CARRERAS_CHOICES, _normalizar_carreras)
        self._bootstrap_tag_field("modalidades", MODALIDADES_CHOICES, _normalizar_modalidades)
        self._bootstrap_tag_field("comisiones", (), _normalizar_comisiones)

    def _bootstrap_tag_field(self, field_name, base_choices, normalizer):
        """Agrega al select los valores actuales aunque no estén en el catálogo base."""
        field = self.fields[field_name]
        base_choices = list(base_choices)
        base_map = {code: label for code, label in base_choices}

        if self.is_bound:
            prefixed_name = self.add_prefix(field_name)
            if hasattr(self.data, "getlist"):
                raw_values = self.data.getlist(prefixed_name)
            else:
                raw_values = self.data.get(prefixed_name, [])
            current_values = normalizer(raw_values)
        else:
            instance_values = getattr(self.instance, field_name, None) if self.instance and self.instance.pk else []
            current_values = instance_values or []
            if current_values:
                self.initial[field_name] = current_values

        extra_choices = [(value, value) for value in current_values if value not in base_map]
        field.choices = base_choices + extra_choices

    def clean_carreras(self):
        return _normalizar_carreras(self.cleaned_data.get("carreras"))

    def clean_modalidades(self):
        return _normalizar_modalidades(self.cleaned_data.get("modalidades"))

    def clean_comisiones(self):
        return _normalizar_comisiones(self.cleaned_data.get("comisiones"))


class CursoIngresoForm(BaseCursoForm):
    class Meta(BaseCursoForm.Meta):
        model = CursoIngreso


@admin.register(CursoIngreso)
class CursoIngresoAdmin(BaseCursoAdmin):
    form = CursoIngresoForm
    actions = ['exportar_cursos_excel', 'exportar_cursos_json']

    @admin.action(description="📥 Exportar cursos a Excel")
    def exportar_cursos_excel(self, request, queryset):
        """Exporta los cursos seleccionados a un archivo Excel."""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        from django.contrib import messages

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cursos de Ingreso"

        # Encabezados
        headers = [
            'Nombre', 'Curso Moodle (Shortname)', 'Carreras',
            'Modalidades', 'Comisiones', 'Activo'
        ]

        # Estilo de encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Escribir datos
        for row_num, curso in enumerate(queryset, 2):
            row_data = [
                curso.nombre,
                curso.curso_moodle,
                ", ".join(curso.carreras) if curso.carreras else "",
                ", ".join(curso.modalidades) if curso.modalidades else "",
                ", ".join(curso.comisiones) if curso.comisiones else "",
                'Sí' if curso.activo else 'No',
            ]

            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"cursos_ingreso_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Guardar workbook en respuesta
        wb.save(response)

        self.message_user(
            request,
            f"✅ Exportados {queryset.count()} cursos a Excel",
            level=messages.SUCCESS
        )

        return response

    @admin.action(description="📥 Exportar cursos a JSON")
    def exportar_cursos_json(self, request, queryset):
        """
        Exporta los cursos seleccionados a un archivo JSON.

        El formato exportado es compatible con la importación posterior.
        Los campos JSONField (carreras, modalidades, comisiones) se exportan
        como listas de strings.
        """
        from django.http import HttpResponse
        from django.contrib import messages
        import json
        from datetime import datetime

        # Crear lista de cursos con validación explícita de tipos
        cursos_data = []
        for curso in queryset:
            # Asegurar que los campos JSONField sean listas
            cursos_data.append({
                'nombre': curso.nombre,
                'curso_moodle': curso.curso_moodle,
                'carreras': list(curso.carreras) if curso.carreras else [],
                'modalidades': list(curso.modalidades) if curso.modalidades else [],
                'comisiones': list(curso.comisiones) if curso.comisiones else [],
                'activo': curso.activo,
            })

        # Crear respuesta JSON con codificación UTF-8 explícita
        json_content = json.dumps(cursos_data, indent=2, ensure_ascii=False)
        response = HttpResponse(
            json_content,
            content_type='application/json; charset=utf-8'
        )
        filename = f"cursos_ingreso_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        self.message_user(
            request,
            f"✅ Exportados {queryset.count()} cursos a JSON ({len(json_content)} bytes)",
            level=messages.SUCCESS
        )

        return response

    def get_urls(self):
        """Agregar URL personalizada para importar JSON."""
        urls = super().get_urls()
        from django.urls import path
        custom_urls = [
            path(
                'importar-json/',
                self.admin_site.admin_view(self.importar_json_view),
                name='cursos_cursoingreso_importar_json',
            ),
        ]
        return custom_urls + urls

    def importar_json_view(self, request):
        """Importa cursos desde un archivo JSON."""
        from django.shortcuts import redirect
        from django.contrib import messages
        import json

        if request.method != 'POST':
            messages.error(request, "❌ Método no permitido")
            return redirect('..')

        if 'json_file' not in request.FILES:
            messages.error(request, "❌ No se seleccionó ningún archivo")
            return redirect('..')

        try:
            json_file = request.FILES['json_file']
            data = json.load(json_file)

            if not isinstance(data, list):
                messages.error(request, "❌ El archivo JSON debe contener una lista de cursos")
                return redirect('..')

            created_count = 0
            updated_count = 0
            error_count = 0

            for curso_data in data:
                try:
                    # Buscar si existe un curso con el mismo shortname
                    curso_moodle = curso_data.get('curso_moodle')
                    if not curso_moodle:
                        error_count += 1
                        continue

                    curso, created = CursoIngreso.objects.update_or_create(
                        curso_moodle=curso_moodle,
                        defaults={
                            'nombre': curso_data.get('nombre', ''),
                            'carreras': curso_data.get('carreras', []),
                            'modalidades': curso_data.get('modalidades', []),
                            'comisiones': curso_data.get('comisiones', []),
                            'activo': curso_data.get('activo', True),
                        }
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    error_count += 1
                    messages.warning(
                        request,
                        f"⚠️ Error al importar curso '{curso_data.get('nombre', 'N/A')}': {str(e)}"
                    )

            # Mensaje de resumen
            messages.success(
                request,
                f"✅ Importación completada: {created_count} creados, {updated_count} actualizados"
            )

            if error_count > 0:
                messages.warning(request, f"⚠️ {error_count} cursos con errores")

        except json.JSONDecodeError as e:
            messages.error(request, f"❌ Error al parsear JSON: {str(e)}")
        except Exception as e:
            messages.error(request, f"❌ Error al importar: {str(e)}")

        return redirect('..')

    def changelist_view(self, request, extra_context=None):
        """Agrega contexto para mostrar botón de importar."""
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context)
