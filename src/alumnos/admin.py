"""
Nombre del Módulo: admin.py

Descripción:
Configuración del admin de Django para la aplicación.

Autor: Carlos Dagorret
Fecha de Creación: 2025-12-29
Última Modificación: 2025-12-29

Licencia: MIT
Copyright (c) 2025 Carlos Dagorret

Permisos:
Se concede permiso, de forma gratuita, a cualquier persona que obtenga una copia
de este software y la documentación asociada (el "Software"), para tratar
en el Software sin restricciones, incluyendo, sin limitación, los derechos
de usar, copiar, modificar, fusionar, publicar, distribuir, sublicenciar
y/o vender copias del Software, y para permitir a las personas a las que
se les proporciona el Software hacerlo, sujeto a las siguientes condiciones:

El aviso de copyright anterior y este aviso de permiso se incluirán en todas
las copias o partes sustanciales del Software.

EL SOFTWARE SE PROPORCIONA "TAL CUAL", SIN GARANTÍA DE NINGÚN TIPO, EXPRESA O
IMPLÍCITA, INCLUYENDO PERO NO LIMITADO A LAS GARANTÍAS DE COMERCIABILIDAD,
IDONEIDAD PARA UN PROPÓSITO PARTICULAR Y NO INFRACCIÓN. EN NINGÚN CASO LOS
AUTORES O TITULARES DE LOS DERECHOS DE AUTOR SERÁN RESPONSABLES DE CUALQUIER
RECLAMO, DAÑO U OTRA RESPONSABILIDAD, YA SEA EN UNA ACCIÓN DE CONTRATO,
AGRAVIO O DE OTRO MODO, QUE SURJA DE, FUERA DE O EN CONEXIÓN CON EL SOFTWARE
O EL USO U OTROS TRATOS EN EL SOFTWARE.
"""
from django.contrib import admin
from django.contrib import messages
from django.shortcuts import redirect
from django.urls import path
from django.utils import timezone
from django.utils.html import format_html
from datetime import timedelta
import logging

from .models import Alumno, Log, Configuracion, Tarea
from .services import ingerir_desde_sial  # función en services.py (archivo)
from .services.teams_service import TeamsService  # clase en services/ (directorio)
from .services.email_service import EmailService  # clase en services/ (directorio)

logger = logging.getLogger(__name__)


def extraer_codigo_error(excepcion_msg, codigo_default='G-001'):
    """
    Extrae el código de error si viene en el formato "T-XXX: mensaje".
    Retorna tupla (codigo_error, mensaje_limpio)
    """
    error_msg = str(excepcion_msg)

    # Lista de códigos válidos
    codigos_validos = [
        'T-001', 'T-002', 'T-003', 'T-004', 'T-005', 'T-006', 'T-007', 'T-008', 'T-009', 'T-999',
        'M-001', 'M-002', 'M-003', 'M-004', 'M-005', 'M-006', 'M-007', 'M-008', 'M-009', 'M-010', 'M-011',
        'E-001', 'E-002', 'E-003', 'E-004', 'E-005', 'E-006', 'E-007',
        'U-001', 'U-002', 'U-003', 'U-004', 'U-005', 'U-006',
        'G-001', 'G-002', 'G-003', 'G-004', 'G-005', 'G-006'
    ]

    # Verificar si el mensaje comienza con un código válido
    if ':' in error_msg:
        posible_codigo = error_msg.split(':')[0].strip()
        if posible_codigo in codigos_validos:
            mensaje_limpio = error_msg.split(':', 1)[1].strip()
            return (posible_codigo, mensaje_limpio)

    return (codigo_default, error_msg)


def encolar_o_ejecutar_tarea(alumno, tipo_tarea, task_func=None, task_args=None, usuario=None, detalles=None):
    """
    Helper para encolar tarea o ejecutarla inmediatamente según USE_QUEUE_SYSTEM.

    Args:
        alumno: Instancia de Alumno
        tipo_tarea: Tarea.TipoTarea (ej: Tarea.TipoTarea.CREAR_USUARIO_TEAMS)
        task_func: Función de tarea de Celery (ej: crear_usuario_teams_async)
        task_args: Tupla con argumentos para task_func (ej: (alumno.id,))
        usuario: Nombre de usuario que ejecuta la acción
        detalles: Dict con detalles adicionales para la tarea

    Returns:
        Objeto Tarea creado
    """
    from django.conf import settings

    use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)

    if use_queue:
        # NUEVO: Solo crear registro Tarea (procesador lo ejecutará)
        tarea = Tarea.objects.create(
            tipo=tipo_tarea,
            estado=Tarea.EstadoTarea.PENDING,
            alumno=alumno,
            usuario=usuario,
            detalles=detalles or {}
        )
    else:
        # LEGACY: Ejecutar inmediatamente con .delay()
        if task_func and task_args is not None:
            task = task_func.delay(*task_args)
            celery_task_id = task.id
        else:
            celery_task_id = None

        tarea = Tarea.objects.create(
            tipo=tipo_tarea,
            estado=Tarea.EstadoTarea.PENDING,
            celery_task_id=celery_task_id,
            alumno=alumno,
            usuario=usuario,
            detalles=detalles or {}
        )

    return tarea


@admin.register(Alumno)
class AlumnoAdmin(admin.ModelAdmin):
    change_list_template = "admin/alumnos/alumno/change_list.html"
    list_display = (
        "apellido",
        "nombre",
        "tipo_documento",
        "dni",
        "estado_actual",
        "modalidad_actual",
        "carreras_display",  # Muestra solo la primera carrera
        "cohorte",
        "fecha_ingreso",
        "teams_status",
        "moodle_status",
        "email_status",
    )
    list_filter = ("estado_actual", "modalidad_actual", "cohorte")
    search_fields = ("apellido", "nombre", "dni", "email_personal", "email_institucional", "carreras_data__nombre_carrera")
    readonly_fields = (
        "created_at",
        "updated_at",
        "fecha_ultima_modificacion",
        "carreras_formatted",
    )
    ordering = ("apellido", "nombre")
    actions = [
        # ===== ACCIONES ATÓMICAS (SÍNCRONAS) =====
        'enviar_email_bienvenida_masivo',
        'activar_teams_con_email_sync',
        'activar_teams_sin_email_sync',
        'resetear_password_teams_con_email_sync',
        'resetear_password_teams_sin_email_sync',
        'borrar_teams_sync',
        'enrollar_moodle_con_email_sync',
        'enrollar_moodle_sin_email_sync',
        'desenrollar_moodle_con_email_sync',
        'desenrollar_moodle_sin_email_sync',
        'borrar_solo_de_moodle',
        'activar_teams_y_moodle_con_email_sync',
        'activar_teams_y_moodle_sin_email_sync',
        # ===== EXPORTACIÓN/IMPORTACIÓN =====
        'exportar_alumnos_excel',
    ]

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "ingesta/",
                self.admin_site.admin_view(self.ingesta_view),
                name="alumnos_alumno_ingesta",
            ),
        ]
        return custom + urls

    def ingesta_view(self, request):
        """
        Endpoint interno para disparar ingestas o borrar alumnos desde la UI del admin.
        Usa POST con `action` para diferenciar.

        IMPORTANTE: La ingesta ahora se ejecuta en la cola de Celery (asíncrona).
        """
        if request.method != "POST":
            return redirect("..")

        action = request.POST.get("action")
        n = request.POST.get("n") or None
        seed = request.POST.get("seed") or None
        desde = request.POST.get("desde") or None
        hasta = request.POST.get("hasta") or None
        tipo = request.POST.get("tipo") or None

        if action == "truncate":
            deleted, _ = Alumno.objects.all().delete()
            messages.success(request, f"Alumnos borrados: {deleted}")
            return redirect("..")

        if action == "seed_pre":
            tipo = "preinscriptos"
        elif action == "seed_asp":
            tipo = "aspirantes"
        elif action == "consume":
            if not tipo:
                messages.error(request, "Debe indicar un tipo (preinscriptos, aspirantes, ingresantes).")
                return redirect("..")
            if (desde and not hasta) or (hasta and not desde):
                messages.error(request, "Si usa rango, debe completar 'desde' y 'hasta'.")
                return redirect("..")
        else:
            messages.error(request, "Accion no reconocida.")
            return redirect("..")

        try:
            n_int = int(n) if n is not None and n != "" else None
        except ValueError:
            messages.error(request, "Cantidad 'n' invalida.")
            return redirect("..")

        try:
            seed_int = int(seed) if seed is not None and seed != "" else None
        except ValueError:
            messages.error(request, "Seed invalida.")
            return redirect("..")

        # Leer checkbox de enviar_email
        enviar_email = request.POST.get("enviar_email") == "1"

        # 🔧 REPARACIÓN: Enviar a cola de Celery en lugar de ejecutar síncronamente
        from .tasks import ingesta_manual_task

        task = ingesta_manual_task.delay(
            tipo=tipo,
            n=n_int,
            seed=seed_int,
            desde=desde,
            hasta=hasta,
            enviar_email=enviar_email,
            usuario=request.user.username if request.user.is_authenticated else None
        )

        # NOTA: La tarea asíncrona crea su propio registro en Tareas Asíncronas
        # No lo creamos aquí para evitar duplicados (IntegrityError)

        messages.success(
            request,
            f"✅ Ingesta de {tipo} programada en cola de Celery. "
            f"Revisa la tabla de Tareas Asíncronas para ver el progreso (ID: {task.id[:8]}...)"
        )
        return redirect("..")

    # =====================================================
    # ACCIONES ATÓMICAS (SÍNCRONAS)
    # =====================================================

    @admin.action(description="⚡ Teams con email (ATÓMICO)")
    def activar_teams_con_email_sync(self, request, queryset):
        """
        Crea usuario en Teams y envía email con credenciales.
        Ejecución SÍNCRONA (sin cola).
        """
        from django.conf import settings
        from .services.teams_service import TeamsService
        from .services.email_service import EmailService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        teams_svc = TeamsService()
        email_svc = EmailService()

        exitos = 0
        errores = 0

        for alumno in queryset:
            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None
            mensaje_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Teams con email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.CREAR_USUARIO_TEAMS,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'activar_teams_con_email',
                    'enviar_email': True
                }
            )

            try:
                # Validar email
                if not alumno.email:
                    codigo_error = 'T-001'
                    raise ValueError("No tiene email configurado")

                # Crear usuario en Teams
                logger.info(f"Llamando teams_svc.create_user para alumno {alumno.id}")
                result = teams_svc.create_user(alumno)
                logger.info(f"Resultado de Teams: type={type(result)}, value={result}")

                if not result:
                    codigo_error = 'T-003'
                    raise Exception("Error al crear usuario en Teams")

                # Validar que el resultado tenga los campos necesarios
                if not isinstance(result, dict):
                    codigo_error = 'T-003'
                    raise Exception(f"Respuesta inválida de Teams - no es dict: {type(result).__name__}, valor: {str(result)[:100]}")

                if 'upn' not in result:
                    codigo_error = 'T-003'
                    raise Exception(f"Respuesta de Teams sin 'upn': {list(result.keys())}")

                # Actualizar alumno
                alumno.email_institucional = result.get('upn')
                alumno.teams_password = result.get('password')
                alumno.teams_procesado = True
                # alumno.teams_payload = result
                alumno.save(update_fields=[
                    'email_institucional', 'teams_password',
                    'teams_procesado'
                ])

                # Enviar email solo si tenemos password
                email_enviado = False
                if result.get('password'):
                    teams_data = {
                        'upn': result.get('upn'),
                        'password': result.get('password')
                    }
                    email_enviado = email_svc.send_credentials_email(alumno, teams_data)
                    if email_enviado:
                        alumno.email_procesado = True
                        alumno.save(update_fields=['email_procesado'])
                else:
                    logger.warning(f"No se pudo enviar email a {alumno}: sin password en resultado de Teams")

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['upn'] = result['upn']
                tarea.detalles['email_enviado'] = email_enviado
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Teams con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'upn': result['upn']
                    }
                )

                exitos += 1

            except Exception as e:
                # Actualizar tarea como fallida
                fin = timezone.now()
                duracion = time.time() - inicio_time

                # Extraer código de error del mensaje si viene en formato "T-XXX: mensaje"
                codigo_final, mensaje_limpio = extraer_codigo_error(e, codigo_error)

                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = mensaje_limpio
                tarea.detalles['codigo_error'] = codigo_final
                tarea.detalles['error'] = mensaje_limpio
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log error
                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Teams con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_final,
                        'error': mensaje_limpio,
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        self.message_user(
            request,
            f"✅ Teams: {exitos} éxitos, {errores} errores. Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="⚡ Teams sin email (ATÓMICO)")
    def activar_teams_sin_email_sync(self, request, queryset):
        """
        Crea usuario en Teams SIN enviar email.
        Ejecución SÍNCRONA (sin cola).
        """
        from django.conf import settings
        from .services.teams_service import TeamsService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        teams_svc = TeamsService()

        exitos = 0
        errores = 0

        for alumno in queryset:
            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Teams sin email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.CREAR_USUARIO_TEAMS,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'activar_teams_sin_email',
                    'enviar_email': False
                }
            )

            try:
                # Crear usuario en Teams
                result = teams_svc.create_user(alumno)

                if not result:
                    codigo_error = 'T-003'
                    raise Exception("Error al crear usuario en Teams")

                # Validar que el resultado tenga los campos necesarios
                if not isinstance(result, dict) or 'upn' not in result:
                    codigo_error = 'T-003'
                    raise Exception(f"Respuesta inválida de Teams: {type(result)}")

                # Actualizar alumno
                alumno.email_institucional = result.get('upn')
                alumno.teams_password = result.get('password')
                alumno.teams_procesado = True
                # alumno.teams_payload = result
                alumno.save(update_fields=[
                    'email_institucional', 'teams_password',
                    'teams_procesado'
                ])

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['upn'] = result['upn']
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Teams sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'upn': result['upn']
                    }
                )

                exitos += 1

            except Exception as e:
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = str(e)
                tarea.detalles['codigo_error'] = codigo_error or 'G-001'
                tarea.detalles['error'] = str(e)
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Teams sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_error or 'G-001',
                        'error': str(e),
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        self.message_user(
            request,
            f"✅ Teams: {exitos} éxitos, {errores} errores. Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="⚡ Moodle con email (ATÓMICO)")
    def enrollar_moodle_con_email_sync(self, request, queryset):
        """
        Enrolla en Moodle y envía email de enrollamiento.
        Ejecución SÍNCRONA (sin cola).
        """
        from .services.moodle_service import MoodleService
        from .services.email_service import EmailService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        moodle_svc = MoodleService()
        email_svc = EmailService()

        exitos = 0
        errores = 0

        for alumno in queryset:
            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Moodle con email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.MOODLE_ENROLL,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'enrollar_moodle_con_email',
                    'enviar_email': True
                }
            )

            try:
                if not alumno.email_institucional:
                    codigo_error = 'M-001'
                    raise ValueError("No tiene email institucional")

                # Crear/buscar usuario en Moodle
                result = moodle_svc.create_user(alumno)
                if not result:
                    codigo_error = 'M-004'
                    raise Exception('Error al crear usuario en Moodle - Sin respuesta')

                if isinstance(result, dict) and 'error' in result:
                    codigo_error = 'M-004'
                    error_msg = result.get('error') if isinstance(result, dict) else str(result)
                    raise Exception(f'Error al crear usuario en Moodle: {error_msg}')

                # Enrollar en cursos
                enroll_result = moodle_svc.enroll_user_in_courses(alumno)
                if not enroll_result:
                    codigo_error = 'M-006'
                    raise Exception('Error al enrollar en cursos - Sin respuesta')

                if isinstance(enroll_result, dict) and 'error' in enroll_result:
                    codigo_error = 'M-006'
                    error_msg = enroll_result.get('error') if isinstance(enroll_result, dict) else str(enroll_result)
                    raise Exception(f'Error al enrollar en cursos: {error_msg}')

                # Actualizar alumno
                alumno.moodle_procesado = True
                # alumno.moodle_payload = {'user': result, 'enrollments': enroll_result}
                alumno.save(update_fields=['moodle_procesado'])

                # Enviar email
                email_enviado = email_svc.send_enrollment_email(alumno)
                if email_enviado:
                    alumno.email_procesado = True
                    alumno.save(update_fields=['email_procesado'])

                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['email_enviado'] = email_enviado
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Moodle con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={'duracion_segundos': round(duracion, 2)}
                )

                exitos += 1

            except Exception as e:
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = str(e)
                tarea.detalles['codigo_error'] = codigo_error or 'G-001'
                tarea.detalles['error'] = str(e)
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Moodle con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_error or 'G-001',
                        'error': str(e),
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        self.message_user(
            request,
            f"✅ Moodle: {exitos} éxitos, {errores} errores. Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="⚡ Moodle sin email (ATÓMICO)")
    def enrollar_moodle_sin_email_sync(self, request, queryset):
        """
        Enrolla en Moodle SIN enviar email.
        Ejecución SÍNCRONA (sin cola).
        """
        from .services.moodle_service import MoodleService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        moodle_svc = MoodleService()

        exitos = 0
        errores = 0

        for alumno in queryset:
            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Moodle sin email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.MOODLE_ENROLL,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'enrollar_moodle_sin_email',
                    'enviar_email': False
                }
            )

            try:
                if not alumno.email_institucional:
                    codigo_error = 'M-001'
                    raise ValueError("No tiene email institucional")

                # Crear/buscar usuario en Moodle
                result = moodle_svc.create_user(alumno)
                if not result:
                    codigo_error = 'M-004'
                    raise Exception('Error al crear usuario en Moodle - Sin respuesta')

                if isinstance(result, dict) and 'error' in result:
                    codigo_error = 'M-004'
                    error_msg = result.get('error') if isinstance(result, dict) else str(result)
                    raise Exception(f'Error al crear usuario en Moodle: {error_msg}')

                # Enrollar en cursos
                enroll_result = moodle_svc.enroll_user_in_courses(alumno)
                if not enroll_result:
                    codigo_error = 'M-006'
                    raise Exception('Error al enrollar en cursos - Sin respuesta')

                if isinstance(enroll_result, dict) and 'error' in enroll_result:
                    codigo_error = 'M-006'
                    error_msg = enroll_result.get('error') if isinstance(enroll_result, dict) else str(enroll_result)
                    raise Exception(f'Error al enrollar en cursos: {error_msg}')

                # Actualizar alumno
                alumno.moodle_procesado = True
                # alumno.moodle_payload = {'user': result, 'enrollments': enroll_result}
                alumno.save(update_fields=['moodle_procesado'])

                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Moodle sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={'duracion_segundos': round(duracion, 2)}
                )

                exitos += 1

            except Exception as e:
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = str(e)
                tarea.detalles['codigo_error'] = codigo_error or 'G-001'
                tarea.detalles['error'] = str(e)
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Moodle sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_error or 'G-001',
                        'error': str(e),
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        self.message_user(
            request,
            f"✅ Moodle: {exitos} éxitos, {errores} errores. Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="⚡⚡ Teams + Moodle con email (ATÓMICO)")
    def activar_teams_y_moodle_con_email_sync(self, request, queryset):
        """
        Crea usuario en Teams, enrolla en Moodle y envía emails.
        Operación ATÓMICA.
        """
        from django.conf import settings
        from .services.teams_service import TeamsService
        from .services.moodle_service import MoodleService
        from .services.email_service import EmailService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        teams_svc = TeamsService()
        moodle_svc = MoodleService()
        email_svc = EmailService()

        exitos = 0
        errores = 0

        for alumno in queryset:
            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None
            paso = ""

            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Teams + Moodle con email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.ACTIVAR_SERVICIOS,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'activar_teams_y_moodle_con_email',
                    'enviar_email': True
                }
            )

            try:
                if not alumno.email:
                    codigo_error = 'T-001'
                    raise ValueError("No tiene email configurado")

                # PASO 1: Crear usuario en Teams
                paso = "Teams"
                teams_result = teams_svc.create_user(alumno)

                if not teams_result:
                    codigo_error = 'T-003'
                    raise Exception("Error al crear usuario en Teams")

                # Validar que el resultado tenga los campos necesarios
                if not isinstance(teams_result, dict) or 'upn' not in teams_result:
                    codigo_error = 'T-003'
                    raise Exception(f"Respuesta inválida de Teams: {type(teams_result)}")

                alumno.email_institucional = teams_result.get('upn')
                alumno.teams_password = teams_result.get('password')
                alumno.teams_procesado = True
                # alumno.teams_payload = teams_result
                alumno.save(update_fields=[
                    'email_institucional', 'teams_password',
                    'teams_procesado'
                ])

                # PASO 2: Enviar email con credenciales Teams (solo si tenemos password)
                if teams_result.get('password'):
                    teams_data = {
                        'upn': teams_result.get('upn'),
                        'password': teams_result.get('password')
                    }
                    email_svc.send_credentials_email(alumno, teams_data)

                # PASO 3: Enrollar en Moodle
                paso = "Moodle"
                moodle_result = moodle_svc.create_user(alumno)
                if not moodle_result or 'error' in moodle_result:
                    codigo_error = 'M-004'
                    raise Exception(moodle_result.get('error', 'Error al crear usuario en Moodle'))

                enroll_result = moodle_svc.enroll_user_in_courses(alumno)
                if not enroll_result or 'error' in enroll_result:
                    codigo_error = 'M-006'
                    raise Exception(enroll_result.get('error', 'Error al enrollar en cursos'))

                alumno.moodle_procesado = True
                # alumno.moodle_payload = {'user': moodle_result, 'enrollments': enroll_result}
                alumno.save(update_fields=['moodle_procesado'])

                # PASO 4: Enviar email de enrollamiento
                email_svc.send_enrollment_email(alumno)
                alumno.email_procesado = True
                alumno.save(update_fields=['email_procesado'])

                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito completo'
                tarea.detalles['upn'] = teams_result['upn']
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Teams + Moodle con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'upn': teams_result['upn']
                    }
                )

                exitos += 1

            except Exception as e:
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = f"Error en {paso}: {str(e)}"
                tarea.detalles['codigo_error'] = codigo_error or 'G-001'
                tarea.detalles['paso_fallido'] = paso
                tarea.detalles['error'] = str(e)
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Teams + Moodle (paso: {paso})',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_error or 'G-001',
                        'paso_fallido': paso,
                        'error': str(e),
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        self.message_user(
            request,
            f"✅ Teams + Moodle: {exitos} éxitos, {errores} errores. Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="⚡⚡ Teams + Moodle sin email (ATÓMICO)")
    def activar_teams_y_moodle_sin_email_sync(self, request, queryset):
        """
        Crea usuario en Teams y enrolla en Moodle SIN enviar emails.
        Operación ATÓMICA.
        """
        from django.conf import settings
        from .services.teams_service import TeamsService
        from .services.moodle_service import MoodleService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        teams_svc = TeamsService()
        moodle_svc = MoodleService()

        exitos = 0
        errores = 0

        for alumno in queryset:
            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None
            paso = ""

            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Teams + Moodle sin email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.ACTIVAR_SERVICIOS,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'activar_teams_y_moodle_sin_email',
                    'enviar_email': False
                }
            )

            try:
                # PASO 1: Crear usuario en Teams
                paso = "Teams"
                teams_result = teams_svc.create_user(alumno)

                if not teams_result:
                    codigo_error = 'T-003'
                    raise Exception("Error al crear usuario en Teams")

                # Validar que el resultado tenga los campos necesarios
                if not isinstance(teams_result, dict) or 'upn' not in teams_result:
                    codigo_error = 'T-003'
                    raise Exception(f"Respuesta inválida de Teams: {type(teams_result)}")

                alumno.email_institucional = teams_result.get('upn')
                alumno.teams_password = teams_result.get('password')
                alumno.teams_procesado = True
                # alumno.teams_payload = teams_result
                alumno.save(update_fields=[
                    'email_institucional', 'teams_password',
                    'teams_procesado'
                ])

                # PASO 2: Enrollar en Moodle
                paso = "Moodle"
                moodle_result = moodle_svc.create_user(alumno)
                if not moodle_result or 'error' in moodle_result:
                    codigo_error = 'M-004'
                    raise Exception(moodle_result.get('error', 'Error al crear usuario en Moodle'))

                enroll_result = moodle_svc.enroll_user_in_courses(alumno)
                if not enroll_result or 'error' in enroll_result:
                    codigo_error = 'M-006'
                    raise Exception(enroll_result.get('error', 'Error al enrollar en cursos'))

                alumno.moodle_procesado = True
                # alumno.moodle_payload = {'user': moodle_result, 'enrollments': enroll_result}
                alumno.save(update_fields=['moodle_procesado'])

                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito completo'
                tarea.detalles['upn'] = teams_result['upn']
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Teams + Moodle sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'upn': teams_result['upn']
                    }
                )

                exitos += 1

            except Exception as e:
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = f"Error en {paso}: {str(e)}"
                tarea.detalles['codigo_error'] = codigo_error or 'G-001'
                tarea.detalles['paso_fallido'] = paso
                tarea.detalles['error'] = str(e)
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Teams + Moodle (paso: {paso})',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_error or 'G-001',
                        'paso_fallido': paso,
                        'error': str(e),
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        self.message_user(
            request,
            f"✅ Teams + Moodle: {exitos} éxitos, {errores} errores. Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    # =====================================================
    # ACCIONES ASÍNCRONAS (CON COLA CELERY)
    # =====================================================

    @admin.action(description="🚀 Activar Teams + Enviar Email con credenciales")
    def activar_teams_email(self, request, queryset):
        """
        Crea usuario en Teams y envía email con credenciales (modo asíncrono).

        **Comportamiento según USE_QUEUE_SYSTEM**:
        - False: Ejecuta inmediatamente con .delay()
        - True: Encola para procesamiento cada 5 min con rate limiting
        """
        from django.conf import settings
        from .tasks import activar_servicios_alumno

        tareas_programadas = 0
        skipped_count = 0
        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        for alumno in queryset:
            # Validar que tenga email
            if not alumno.email:
                self.message_user(
                    request,
                    f"⚠️ {alumno.apellido}, {alumno.nombre} no tiene email configurado",
                    level=messages.WARNING
                )
                skipped_count += 1
                continue

            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.ACTIVAR_SERVICIOS,
                task_func=activar_servicios_alumno,
                task_args=(alumno.id,),
                usuario=usuario
            )
            tareas_programadas += 1

        # Mensaje según modo
        if use_queue:
            mensaje = f"✅ {tareas_programadas} tareas encoladas. Serán procesadas en máx 5 minutos."
        else:
            mensaje = f"📋 {tareas_programadas} tareas ejecutándose en background."

        self.message_user(request, mensaje + " Ver Tareas Asíncronas.", level=messages.SUCCESS)

        if skipped_count > 0:
            self.message_user(
                request,
                f"⚠️ {skipped_count} alumnos omitidos por falta de email",
                level=messages.WARNING
            )

    @admin.action(description="🎓 Enrollar en Moodle + Email de Enrollamiento")
    def enrollar_moodle_con_email(self, request, queryset):
        """
        Enrolla alumnos en Moodle y envía email de enrollamiento (Ecosistema Virtual).
        """
        from django.conf import settings
        from .tasks import enrollar_moodle_task
        from .models import Tarea

        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        tareas_programadas = 0
        already_processed = 0
        skipped_count = 0

        for alumno in queryset:
            # Validar que tenga email
            if not alumno.email:
                self.message_user(
                    request,
                    f"⚠️ {alumno.apellido}, {alumno.nombre} no tiene email configurado",
                    level=messages.WARNING
                )
                skipped_count += 1
                continue

            # Verificar si ya está procesado en Moodle
            if alumno.moodle_procesado:
                already_processed += 1
                continue

            # Encolar o ejecutar tarea con envío de email
            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.MOODLE_ENROLL,
                task_func=enrollar_moodle_task,
                task_args=(alumno.id, True),  # enviar_email=True
                usuario=usuario,
                detalles={'enviar_email': True}
            )

            tareas_programadas += 1

        # Resumen final
        if tareas_programadas > 0:
            modo = "encoladas" if use_queue else "programadas"
            self.message_user(
                request,
                f"📋 {tareas_programadas} tareas {modo} para enrollamiento en Moodle (con email de enrollamiento).",
                level=messages.SUCCESS
            )

        if already_processed > 0:
            self.message_user(
                request,
                f"✅ {already_processed} alumnos ya enrollados en Moodle (omitidos)",
                level=messages.INFO
            )

        if skipped_count > 0:
            self.message_user(
                request,
                f"⚠️ {skipped_count} alumnos omitidos por falta de email",
                level=messages.WARNING
            )

    @admin.action(description="🎓 Enrollar en Moodle (sin email)")
    def enrollar_moodle_sin_email(self, request, queryset):
        """
        Enrolla alumnos en Moodle SIN enviar email de bienvenida.
        """
        from django.conf import settings
        from .tasks import enrollar_moodle_task
        from .models import Tarea

        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        tareas_programadas = 0
        already_processed = 0
        skipped_count = 0

        for alumno in queryset:
            # Validar que tenga email institucional
            if not alumno.email_institucional:
                self.message_user(
                    request,
                    f"⚠️ {alumno.apellido}, {alumno.nombre} no tiene email institucional",
                    level=messages.WARNING
                )
                skipped_count += 1
                continue

            # Verificar si ya está procesado en Moodle
            if alumno.moodle_procesado:
                already_processed += 1
                continue

            # Encolar o ejecutar tarea SIN envío de email
            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.MOODLE_ENROLL,
                task_func=enrollar_moodle_task,
                task_args=(alumno.id, False),  # enviar_email=False
                usuario=usuario,
                detalles={'enviar_email': False}
            )

            tareas_programadas += 1

        # Resumen final
        if tareas_programadas > 0:
            modo = "encoladas" if use_queue else "programadas"
            self.message_user(
                request,
                f"📋 {tareas_programadas} tareas {modo} para enrollamiento en Moodle (sin email).",
                level=messages.SUCCESS
            )

        if already_processed > 0:
            self.message_user(
                request,
                f"✅ {already_processed} alumnos ya enrollados en Moodle (omitidos)",
                level=messages.INFO
            )

        if skipped_count > 0:
            self.message_user(
                request,
                f"⚠️ {skipped_count} alumnos omitidos por falta de email institucional",
                level=messages.WARNING
            )

    @admin.action(description="📧 Enviar email de bienvenida (ATÓMICO)")
    def enviar_email_bienvenida_masivo(self, request, queryset):
        """
        Envía email de bienvenida a los alumnos seleccionados.
        Ejecución SÍNCRONA con registro en Tareas Asíncronas.
        """
        from .services.email_service import EmailService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        email_svc = EmailService()
        exitos = 0
        errores = 0

        for alumno in queryset:
            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Enviar email bienvenida',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.ENVIAR_EMAIL,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'enviar_email_bienvenida',
                    'tipo_email': 'bienvenida'
                }
            )

            try:
                # Validar que tenga email
                if not alumno.email:
                    codigo_error = 'E-001'
                    raise ValueError("No tiene email personal configurado")

                # Enviar email
                result = email_svc.send_welcome_email(alumno)

                if not result:
                    codigo_error = 'E-002'
                    raise Exception("Error al enviar email de bienvenida")

                # Actualizar alumno - MARCAR COMO PROCESADO
                alumno.email_procesado = True
                alumno.save(update_fields=['email_procesado'])

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['email_enviado_a'] = alumno.email
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Enviar email bienvenida',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'email': alumno.email
                    }
                )

                exitos += 1

            except Exception as e:
                # Actualizar tarea como fallida
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = str(e)
                tarea.detalles['codigo_error'] = codigo_error or 'G-001'
                tarea.detalles['error'] = str(e)
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log error
                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Enviar email bienvenida',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_error or 'G-001',
                        'error': str(e),
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        self.message_user(
            request,
            f"✅ Emails: {exitos} enviados, {errores} errores. Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="🗑️ Borrar de Teams (ATÓMICO)")
    def borrar_teams_sync(self, request, queryset):
        """
        Elimina usuarios de Teams de forma SÍNCRONA.
        PRECAUCIÓN: Solo funciona con cuentas test-*
        """
        from .services.teams_service import TeamsService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        teams_svc = TeamsService()
        exitos = 0
        errores = 0
        omitidos = 0

        for alumno in queryset:
            # Validar que tenga email institucional
            if not alumno.email_institucional:
                omitidos += 1
                continue

            # Validar que esté procesado en Teams
            if not alumno.teams_procesado:
                omitidos += 1
                continue

            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Borrar de Teams',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.ELIMINAR_CUENTA,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'borrar_teams',
                    'servicio': 'Teams',
                    'upn': alumno.email_institucional
                }
            )

            try:
                upn = alumno.email_institucional

                # Eliminar usuario de Teams
                teams_svc.delete_user(upn, alumno)

                # Limpiar campos del alumno
                alumno.email_institucional = None
                alumno.teams_password = None
                alumno.teams_procesado = False
                # alumno.teams_payload = None
                alumno.save(update_fields=[
                    'email_institucional', 'teams_password',
                    'teams_procesado'
                ])

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['upn_eliminado'] = upn
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Borrar de Teams',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'upn': upn
                    }
                )

                exitos += 1

            except Exception as e:
                # Actualizar tarea como fallida
                fin = timezone.now()
                duracion = time.time() - inicio_time

                # Extraer código de error del mensaje si viene en formato "T-XXX: mensaje"
                codigo_final, mensaje_limpio = extraer_codigo_error(e, codigo_error)

                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = mensaje_limpio
                tarea.detalles['codigo_error'] = codigo_final
                tarea.detalles['error'] = mensaje_limpio
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log error
                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Borrar de Teams',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_final,
                        'error': mensaje_limpio,
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        mensaje = f"🗑️ Teams: {exitos} eliminados, {errores} errores"
        if omitidos > 0:
            mensaje += f", {omitidos} omitidos"

        self.message_user(
            request,
            mensaje + ". Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="🔄 Resetear password Teams con email (ATÓMICO)")
    def resetear_password_teams_con_email_sync(self, request, queryset):
        """
        Resetea contraseña de Teams y envía email con credenciales.
        Ejecución SÍNCRONA con registro en Tareas Asíncronas.
        """
        from .services.teams_service import TeamsService
        from .services.email_service import EmailService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        teams_svc = TeamsService()
        email_svc = EmailService()
        exitos = 0
        errores = 0
        omitidos = 0

        for alumno in queryset:
            # Validar que tenga email institucional
            if not alumno.email_institucional:
                omitidos += 1
                continue

            # Validar que tenga email personal
            if not alumno.email:
                omitidos += 1
                continue

            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Resetear password Teams con email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.RESETEAR_PASSWORD,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'resetear_password_teams_con_email',
                    'upn': alumno.email_institucional,
                    'enviar_email': True
                }
            )

            try:
                upn = alumno.email_institucional

                # Resetear contraseña
                new_password = teams_svc.reset_password(upn, alumno=alumno)

                # Actualizar alumno
                alumno.teams_password = new_password
                alumno.save(update_fields=['teams_password'])

                # Enviar email con nueva contraseña
                teams_data = {
                    'upn': upn,
                    'password': new_password
                }
                email_enviado = email_svc.send_credentials_email(alumno, teams_data)

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['upn'] = upn
                tarea.detalles['email_enviado'] = email_enviado
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Resetear password Teams con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'upn': upn
                    }
                )

                exitos += 1

            except Exception as e:
                # Actualizar tarea como fallida
                fin = timezone.now()
                duracion = time.time() - inicio_time

                # Extraer código de error del mensaje si viene en formato "T-XXX: mensaje"
                codigo_final, mensaje_limpio = extraer_codigo_error(e, codigo_error)

                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = mensaje_limpio
                tarea.detalles['codigo_error'] = codigo_final
                tarea.detalles['error'] = mensaje_limpio
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log error
                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Resetear password Teams con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_final,
                        'error': mensaje_limpio,
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        mensaje = f"🔄 Password Teams: {exitos} reseteados, {errores} errores"
        if omitidos > 0:
            mensaje += f", {omitidos} omitidos"

        self.message_user(
            request,
            mensaje + ". Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="🔄 Resetear password Teams sin email (ATÓMICO)")
    def resetear_password_teams_sin_email_sync(self, request, queryset):
        """
        Resetea contraseña de Teams SIN enviar email.
        Ejecución SÍNCRONA con registro en Tareas Asíncronas.
        """
        from .services.teams_service import TeamsService
        from .models import Log, Tarea
        from django.utils import timezone
        import time

        teams_svc = TeamsService()
        exitos = 0
        errores = 0
        omitidos = 0

        for alumno in queryset:
            # Validar que tenga email institucional
            if not alumno.email_institucional:
                omitidos += 1
                continue

            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Resetear password Teams sin email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.RESETEAR_PASSWORD,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'resetear_password_teams_sin_email',
                    'upn': alumno.email_institucional,
                    'enviar_email': False
                }
            )

            try:
                upn = alumno.email_institucional

                # Resetear contraseña
                new_password = teams_svc.reset_password(upn, alumno=alumno)

                # Actualizar alumno
                alumno.teams_password = new_password
                alumno.save(update_fields=['teams_password'])

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = 1
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['upn'] = upn
                tarea.detalles['new_password'] = new_password
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Resetear password Teams sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'upn': upn
                    }
                )

                exitos += 1

            except Exception as e:
                # Actualizar tarea como fallida
                fin = timezone.now()
                duracion = time.time() - inicio_time

                # Extraer código de error del mensaje si viene en formato "T-XXX: mensaje"
                codigo_final, mensaje_limpio = extraer_codigo_error(e, codigo_error)

                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = mensaje_limpio
                tarea.detalles['codigo_error'] = codigo_final
                tarea.detalles['error'] = mensaje_limpio
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log error
                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Resetear password Teams sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_final,
                        'error': mensaje_limpio,
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        mensaje = f"🔄 Password Teams: {exitos} reseteados, {errores} errores"
        if omitidos > 0:
            mensaje += f", {omitidos} omitidos"

        self.message_user(
            request,
            mensaje + ". Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="🔻 Desenrollar de Moodle con email (ATÓMICO)")
    def desenrollar_moodle_con_email_sync(self, request, queryset):
        """
        Des-enrolla de cursos de Moodle y envía email de notificación.
        Ejecución SÍNCRONA con registro en Tareas Asíncronas.
        """
        from .services.moodle_service import MoodleService
        from .services.email_service import EmailService
        from .models import Log, Tarea, Configuracion
        from django.utils import timezone
        import time

        from cursos.services import resolver_curso
        from cursos.constants import CARRERAS_DICT

        config = Configuracion.load()
        moodle_svc = MoodleService()
        email_svc = EmailService()
        exitos = 0
        errores = 0
        omitidos = 0

        for alumno in queryset:
            # Validar email institucional con manejo de fallback
            username = alumno.email_institucional
            if not username:
                if config.deshabilitar_fallback_email_personal:
                    # Fallback deshabilitado, omitir alumno
                    Log.objects.create(
                        tipo='WARNING',
                        modulo='admin_action_sync',
                        mensaje=f'⚠️ FALTA EMAIL INSTITUCIONAL - Alumno {alumno.id} omitido (fallback deshabilitado)',
                        alumno=alumno,
                        usuario=request.user.username if request.user.is_authenticated else None,
                        detalles={'dni': alumno.dni, 'nombre': f'{alumno.nombre} {alumno.apellido}'}
                    )
                    omitidos += 1
                    continue
                else:
                    # Usar fallback a email personal
                    username = alumno.email_personal

            if not username:
                omitidos += 1
                continue

            # Validar que tenga email personal para notificación
            if not alumno.email:
                omitidos += 1
                continue

            # Obtener cursos dinámicamente basado en carreras_data
            courses_to_unenrol = []
            if alumno.carreras_data and isinstance(alumno.carreras_data, list):
                for carrera_data in alumno.carreras_data:
                    id_carrera = carrera_data.get('id_carrera')
                    modalidad = carrera_data.get('modalidad', '').strip()
                    comisiones = carrera_data.get('comisiones', [])
                    comision = comisiones[0].get('nombre_comision', '') if comisiones else ''

                    # Mapear id_carrera a código
                    codigo_carrera = CARRERAS_DICT.get(str(id_carrera)) or CARRERAS_DICT.get(id_carrera)
                    if not codigo_carrera:
                        continue

                    try:
                        cursos = resolver_curso(codigo_carrera, modalidad, comision)
                        courses_to_unenrol.extend(cursos)
                    except Exception:
                        continue

            # Eliminar duplicados
            courses_to_unenrol = list(set(courses_to_unenrol))

            if not courses_to_unenrol:
                omitidos += 1
                continue

            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Desenrollar de Moodle con email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.MOODLE_ENROLL,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'desenrollar_moodle_con_email',
                    'username': username,
                    'cursos': courses_to_unenrol,
                    'enviar_email': True
                }
            )

            try:
                # Buscar usuario en Moodle
                user = moodle_svc.get_user_by_username(username)
                if not user:
                    codigo_error = 'M-001'
                    raise ValueError(f"Usuario no encontrado en Moodle: {username}")

                user_id = user['id']

                # Des-enrollar de cada curso
                desenrollados = []
                for course_shortname in courses_to_unenrol:
                    try:
                        if moodle_svc.unenrol_user_from_course(user_id, course_shortname, alumno):
                            desenrollados.append(course_shortname)
                    except Exception as e:
                        logger.warning(f"Error des-enrollando de {course_shortname}: {e}")

                if not desenrollados:
                    codigo_error = 'M-006'
                    raise Exception("No se pudo des-enrollar de ningún curso")

                # Enviar email de notificación
                # TODO: Implementar plantilla de email para des-enrollamiento
                email_enviado = False  # Por ahora no enviamos email

                # Actualizar alumno
                alumno.moodle_procesado = False
                alumno.save(update_fields=['moodle_procesado'])

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = len(desenrollados)
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['cursos_desenrollados'] = desenrollados
                tarea.detalles['email_enviado'] = email_enviado
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Desenrollar de Moodle con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'cursos': desenrollados
                    }
                )

                exitos += 1

            except Exception as e:
                # Actualizar tarea como fallida
                fin = timezone.now()
                duracion = time.time() - inicio_time

                # Extraer código de error del mensaje si viene en formato "M-XXX: mensaje"
                codigo_final, mensaje_limpio = extraer_codigo_error(e, codigo_error)

                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = mensaje_limpio
                tarea.detalles['codigo_error'] = codigo_final
                tarea.detalles['error'] = mensaje_limpio
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log error
                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Desenrollar de Moodle con email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_final,
                        'error': mensaje_limpio,
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        mensaje = f"🔻 Moodle: {exitos} des-enrollados, {errores} errores"
        if omitidos > 0:
            mensaje += f", {omitidos} omitidos"

        self.message_user(
            request,
            mensaje + ". Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="🔻 Desenrollar de Moodle sin email (ATÓMICO)")
    def desenrollar_moodle_sin_email_sync(self, request, queryset):
        """
        Des-enrolla de cursos de Moodle SIN enviar email.
        Ejecución SÍNCRONA con registro en Tareas Asíncronas.
        """
        from .services.moodle_service import MoodleService
        from .models import Log, Tarea, Configuracion
        from django.utils import timezone
        import time
        from cursos.services import resolver_curso
        from cursos.constants import CARRERAS_DICT

        config = Configuracion.load()
        moodle_svc = MoodleService()
        exitos = 0
        errores = 0
        omitidos = 0

        for alumno in queryset:
            # Validar email institucional con manejo de fallback
            username = alumno.email_institucional
            if not username:
                if config.deshabilitar_fallback_email_personal:
                    # Fallback deshabilitado, omitir alumno
                    Log.objects.create(
                        tipo='WARNING',
                        modulo='admin_action_sync',
                        mensaje=f'⚠️ FALTA EMAIL INSTITUCIONAL - Alumno {alumno.id} omitido (fallback deshabilitado)',
                        alumno=alumno,
                        usuario=request.user.username if request.user.is_authenticated else None,
                        detalles={'dni': alumno.dni, 'nombre': f'{alumno.nombre} {alumno.apellido}'}
                    )
                    omitidos += 1
                    continue
                else:
                    # Usar fallback a email personal
                    username = alumno.email_personal

            if not username:
                omitidos += 1
                continue

            # Obtener cursos dinámicamente basado en carreras_data
            courses_to_unenrol = []
            if alumno.carreras_data and isinstance(alumno.carreras_data, list):
                for carrera_data in alumno.carreras_data:
                    id_carrera = carrera_data.get('id_carrera')
                    modalidad = carrera_data.get('modalidad', '').strip()
                    comisiones = carrera_data.get('comisiones', [])
                    comision = comisiones[0].get('nombre_comision', '') if comisiones else ''

                    # Mapear id_carrera a código
                    codigo_carrera = CARRERAS_DICT.get(str(id_carrera)) or CARRERAS_DICT.get(id_carrera)
                    if not codigo_carrera:
                        continue

                    try:
                        cursos = resolver_curso(codigo_carrera, modalidad, comision)
                        courses_to_unenrol.extend(cursos)
                    except Exception:
                        continue

            # Eliminar duplicados
            courses_to_unenrol = list(set(courses_to_unenrol))

            if not courses_to_unenrol:
                omitidos += 1
                continue

            inicio = timezone.now()
            inicio_time = time.time()
            codigo_error = None

            # Log inicio
            Log.objects.create(
                tipo='INFO',
                modulo='admin_action_sync',
                mensaje=f'Iniciando: Desenrollar de Moodle sin email',
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None
            )

            # Crear tarea
            tarea = Tarea.objects.create(
                tipo=Tarea.TipoTarea.MOODLE_ENROLL,
                estado=Tarea.EstadoTarea.RUNNING,
                alumno=alumno,
                usuario=request.user.username if request.user.is_authenticated else None,
                hora_inicio=inicio,
                detalles={
                    'modulo': 'admin_action_sync',
                    'accion': 'desenrollar_moodle_sin_email',
                    'username': username,
                    'cursos': courses_to_unenrol,
                    'enviar_email': False
                }
            )

            try:
                # Buscar usuario en Moodle
                user = moodle_svc.get_user_by_username(username)
                if not user:
                    codigo_error = 'M-001'
                    raise ValueError(f"Usuario no encontrado en Moodle: {username}")

                user_id = user['id']

                # Des-enrollar de cada curso
                desenrollados = []
                for course_shortname in courses_to_unenrol:
                    try:
                        if moodle_svc.unenrol_user_from_course(user_id, course_shortname, alumno):
                            desenrollados.append(course_shortname)
                    except Exception as e:
                        logger.warning(f"Error des-enrollando de {course_shortname}: {e}")

                if not desenrollados:
                    codigo_error = 'M-006'
                    raise Exception("No se pudo des-enrollar de ningún curso")

                # Actualizar alumno
                alumno.moodle_procesado = False
                alumno.save(update_fields=['moodle_procesado'])

                # Actualizar tarea
                fin = timezone.now()
                duracion = time.time() - inicio_time
                tarea.estado = Tarea.EstadoTarea.COMPLETED
                tarea.hora_fin = fin
                tarea.cantidad_entidades = len(desenrollados)
                tarea.detalles['resultado'] = 'Éxito'
                tarea.detalles['cursos_desenrollados'] = desenrollados
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log fin
                Log.objects.create(
                    tipo='SUCCESS',
                    modulo='admin_action_sync',
                    mensaje=f'Completado: Desenrollar de Moodle sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'duracion_segundos': round(duracion, 2),
                        'cursos': desenrollados
                    }
                )

                exitos += 1

            except Exception as e:
                # Actualizar tarea como fallida
                fin = timezone.now()
                duracion = time.time() - inicio_time

                # Extraer código de error del mensaje si viene en formato "M-XXX: mensaje"
                codigo_final, mensaje_limpio = extraer_codigo_error(e, codigo_error)

                tarea.estado = Tarea.EstadoTarea.FAILED
                tarea.hora_fin = fin
                tarea.mensaje_error = mensaje_limpio
                tarea.detalles['codigo_error'] = codigo_final
                tarea.detalles['error'] = mensaje_limpio
                tarea.detalles['duracion_segundos'] = round(duracion, 2)
                tarea.save()

                # Log error
                Log.objects.create(
                    tipo='ERROR',
                    modulo='admin_action_sync',
                    mensaje=f'Error: Desenrollar de Moodle sin email',
                    alumno=alumno,
                    usuario=request.user.username if request.user.is_authenticated else None,
                    detalles={
                        'codigo_error': codigo_final,
                        'error': mensaje_limpio,
                        'duracion_segundos': round(duracion, 2)
                    }
                )

                errores += 1

        mensaje = f"🔻 Moodle: {exitos} des-enrollados, {errores} errores"
        if omitidos > 0:
            mensaje += f", {omitidos} omitidos"

        self.message_user(
            request,
            mensaje + ". Ver Tareas Asíncronas para detalles.",
            level=messages.SUCCESS if exitos > 0 else messages.ERROR
        )

    @admin.action(description="🗑️ Borrar solo de Teams")
    def borrar_solo_de_teams(self, request, queryset):
        """
        Elimina usuarios solo de Teams (no de Moodle).
        """
        from django.conf import settings
        from .tasks_delete import eliminar_solo_teams
        from .models import Tarea

        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        programadas = 0
        skipped = 0

        for alumno in queryset:
            if not alumno.email_institucional:
                skipped += 1
                continue

            if not alumno.teams_procesado:
                skipped += 1
                continue

            # Encolar o ejecutar tarea de borrado de Teams
            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.ELIMINAR_TEAMS,
                task_func=eliminar_solo_teams,
                task_args=(alumno.id, alumno.email_institucional),
                usuario=usuario
            )
            programadas += 1

        modo = "encoladas" if use_queue else "programadas"
        self.message_user(
            request,
            f"🗑️ {programadas} tareas de borrado de Teams {modo}",
            level=messages.SUCCESS
        )

        if skipped > 0:
            self.message_user(
                request,
                f"⚠️ {skipped} alumnos omitidos (sin email o no procesados en Teams)",
                level=messages.WARNING
            )

    @admin.action(description="🗑️ Borrar solo de Moodle (NO Teams)")
    def borrar_solo_de_moodle(self, request, queryset):
        """
        Elimina usuarios solo de Moodle (no de Teams).
        Intenta borrar sin validar moodle_procesado. Ignora errores de "usuario no encontrado".
        """
        from django.conf import settings
        from .tasks_delete import eliminar_solo_moodle
        from .models import Tarea

        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        programadas = 0
        skipped = 0

        for alumno in queryset:
            if not alumno.email_institucional:
                skipped += 1
                continue

            # Encolar o ejecutar tarea de borrado de Moodle
            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.ELIMINAR_MOODLE,
                task_func=eliminar_solo_moodle,
                task_args=(alumno.id, alumno.email_institucional),
                usuario=usuario
            )
            programadas += 1

        modo = "encoladas" if use_queue else "programadas"
        self.message_user(
            request,
            f"🗑️ {programadas} tareas de borrado de Moodle {modo}",
            level=messages.SUCCESS
        )

        if skipped > 0:
            self.message_user(
                request,
                f"⚠️ {skipped} alumnos omitidos (sin email o no procesados en Moodle)",
                level=messages.WARNING
            )

    @admin.action(description="🗑️ Eliminar solo de BD (sin tareas asíncronas)")
    def eliminar_solo_bd_directo(self, request, queryset):
        """
        Elimina alumnos SOLO de la base de datos, SIN disparar tareas asíncronas.
        NO elimina de Teams ni Moodle.
        PRECAUCIÓN: Esta acción es irreversible y directa.
        """
        total = queryset.count()
        queryset.delete()

        self.message_user(
            request,
            f"🗑️ {total} alumnos eliminados directamente de la base de datos (sin tareas asíncronas).",
            level=messages.SUCCESS
        )

    @admin.action(description="🗑️💀 Eliminar alumno completamente (Teams + Moodle + BD)")
    def eliminar_alumno_completo(self, request, queryset):
        """
        Elimina alumnos completamente: Teams + Moodle + Base de datos.
        PRECAUCIÓN: Esta acción es irreversible.
        """
        from django.conf import settings
        from .tasks_delete import eliminar_alumno_completo as task_eliminar_completo
        from .models import Tarea

        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        programadas = 0

        for alumno in queryset:
            # Encolar o ejecutar tarea de eliminación completa
            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.ELIMINAR_COMPLETO,
                task_func=task_eliminar_completo,
                task_args=(alumno.id,),
                usuario=usuario
            )
            programadas += 1

        modo = "encoladas" if use_queue else "programadas"
        self.message_user(
            request,
            f"🗑️💀 {programadas} tareas de eliminación completa {modo}. "
            f"Los alumnos serán eliminados de Teams, Moodle y la base de datos.",
            level=messages.WARNING
        )

    @admin.action(description="🔄 Generar contraseña y enviar correo")
    def resetear_y_enviar_email(self, request, queryset):
        """
        Resetea contraseña en Teams y envía email (modo asíncrono).
        Las tareas se ejecutan en background vía Celery.
        """
        from django.conf import settings
        from .tasks import resetear_password_y_enviar_email
        from .models import Tarea

        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        tareas_programadas = 0
        skipped_count = 0

        for alumno in queryset:
            if not alumno.email:
                self.message_user(
                    request,
                    f"⚠️ {alumno.apellido}, {alumno.nombre} no tiene email configurado",
                    level=messages.WARNING
                )
                skipped_count += 1
                continue

            # Encolar o ejecutar tarea de reseteo de contraseña
            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.RESETEAR_PASSWORD,
                task_func=resetear_password_y_enviar_email,
                task_args=(alumno.id,),
                usuario=usuario
            )

            tareas_programadas += 1

        # Resumen final
        modo = "encoladas" if use_queue else "programadas"
        self.message_user(
            request,
            f"📋 {tareas_programadas} tareas {modo} en cola de Celery. "
            f"Revisa la tabla de Tareas Asíncronas para ver el progreso.",
            level=messages.SUCCESS
        )

        if skipped_count > 0:
            self.message_user(
                request,
                f"⚠️ {skipped_count} alumnos omitidos por falta de email",
                level=messages.WARNING
            )

    @admin.action(description="👤 Crear usuario en Teams (sin email)")
    def crear_usuario_teams(self, request, queryset):
        """
        Crea usuario en Teams sin enviar email (modo asíncrono).

        **Comportamiento según USE_QUEUE_SYSTEM**:
        - False (default): Ejecuta inmediatamente con .delay() (legacy)
        - True: Encola tareas que se procesarán cada 5 min con rate limiting
        """
        from django.conf import settings
        from .tasks import crear_usuario_teams_async

        tareas_programadas = 0
        use_queue = getattr(settings, 'USE_QUEUE_SYSTEM', False)
        usuario = request.user.username if request.user.is_authenticated else None

        for alumno in queryset:
            encolar_o_ejecutar_tarea(
                alumno=alumno,
                tipo_tarea=Tarea.TipoTarea.CREAR_USUARIO_TEAMS,
                task_func=crear_usuario_teams_async,
                task_args=(alumno.id,),
                usuario=usuario
            )
            tareas_programadas += 1

        # Mensaje según modo
        if use_queue:
            self.message_user(
                request,
                f"✅ {tareas_programadas} tareas encoladas. "
                f"Serán procesadas en máx 5 minutos con rate limiting. "
                f"Revisa Tareas Asíncronas para ver el progreso.",
                level=messages.SUCCESS
            )
        else:
            self.message_user(
                request,
                f"📋 {tareas_programadas} tareas ejecutándose en background. "
                f"Revisa la tabla de Tareas Asíncronas para ver credenciales.",
                level=messages.SUCCESS
            )

    @admin.action(description="📧 Enviar email con credenciales")
    def enviar_email_credenciales(self, request, queryset):
        """
        Envía email con credenciales existentes.
        Útil para reenviar credenciales.
        """
        from django.conf import settings

        email_svc = EmailService()
        teams_svc = TeamsService()

        success_count = 0
        error_count = 0

        for alumno in queryset:
            if not alumno.email:
                self.message_user(
                    request,
                    f"⚠️ {alumno.apellido}, {alumno.nombre} no tiene email",
                    level=messages.WARNING
                )
                error_count += 1
                continue

            try:
                # Construir UPN y obtener usuario de Teams
                prefix = settings.ACCOUNT_PREFIX
                upn = f"{prefix}{alumno.dni}@{settings.TEAMS_DOMAIN}"

                # Verificar que el usuario existe en Teams
                user = teams_svc.get_user(upn)
                if not user:
                    self.message_user(
                        request,
                        f"⚠️ {alumno.apellido}, {alumno.nombre}: Usuario {upn} no existe en Teams",
                        level=messages.WARNING
                    )
                    error_count += 1
                    continue

                # Generar nueva contraseña temporal
                new_password = teams_svc.reset_password(upn)
                if not new_password:
                    self.message_user(
                        request,
                        f"❌ Error reseteando contraseña para {upn}",
                        level=messages.ERROR
                    )
                    error_count += 1
                    continue

                # Enviar email con nueva contraseña
                teams_data = {
                    'upn': upn,
                    'password': new_password
                }

                email_sent = email_svc.send_credentials_email(alumno, teams_data)

                if email_sent:
                    self.message_user(
                        request,
                        f"✅ {alumno.apellido}, {alumno.nombre}: Email enviado a {alumno.email}",
                        level=messages.SUCCESS
                    )
                    success_count += 1
                else:
                    self.message_user(
                        request,
                        f"❌ Error enviando email a {alumno.email}",
                        level=messages.ERROR
                    )
                    error_count += 1

            except Exception as e:
                self.message_user(
                    request,
                    f"❌ Error procesando {alumno.apellido}, {alumno.nombre}: {str(e)}",
                    level=messages.ERROR
                )
                error_count += 1

        self.message_user(
            request,
            f"📊 Emails: {success_count} enviados, {error_count} errores",
            level=messages.INFO
        )

    @admin.action(description="🔑 Resetear contraseña Teams")
    def resetear_password_teams(self, request, queryset):
        """
        Resetea la contraseña de Teams para los usuarios seleccionados.
        """
        from django.conf import settings

        teams_svc = TeamsService()

        success_count = 0
        error_count = 0

        for alumno in queryset:
            try:
                prefix = settings.ACCOUNT_PREFIX
                upn = f"{prefix}{alumno.dni}@{settings.TEAMS_DOMAIN}"

                new_password = teams_svc.reset_password(upn)

                if new_password:
                    # Actualizar campos del alumno
                    alumno.email_institucional = upn
                    alumno.teams_password = new_password
                    alumno.save(update_fields=['email_institucional', 'teams_password'])

                    self.message_user(
                        request,
                        f"✅ {alumno.apellido}, {alumno.nombre} - UPN: {upn} - Nueva pass: {new_password}",
                        level=messages.SUCCESS
                    )
                    success_count += 1
                else:
                    self.message_user(
                        request,
                        f"❌ Error reseteando contraseña para {upn}",
                        level=messages.ERROR
                    )
                    error_count += 1

            except Exception as e:
                self.message_user(
                    request,
                    f"❌ Error procesando {alumno.apellido}, {alumno.nombre}: {str(e)}",
                    level=messages.ERROR
                )
                error_count += 1

        self.message_user(
            request,
            f"📊 Contraseñas: {success_count} reseteadas, {error_count} errores",
            level=messages.INFO
        )

    def carreras_display(self, obj):
        """Muestra la carrera del alumno (siempre la primera)."""
        if not obj.carreras_data or len(obj.carreras_data) == 0:
            return "-"

        # Solo mostrar la primera carrera (caso normal)
        carrera = obj.carreras_data[0]
        id_carrera = carrera.get("id_carrera")
        modalidad = carrera.get("modalidad", "")

        # Obtener nombre real de la carrera desde la base de datos
        from cursos.utils import get_carrera_nombre
        nombre = get_carrera_nombre(id_carrera) if id_carrera else carrera.get("nombre_carrera", "Sin nombre")

        mod_texto = "Presencial" if modalidad == "1" else "Distancia" if modalidad == "2" else modalidad

        return f"{nombre} ({mod_texto})"

    carreras_display.short_description = "Carrera"

    def teams_status(self, obj):
        """Muestra estado de Teams con emoticono."""
        from django.utils.safestring import mark_safe
        if obj.teams_procesado:
            return mark_safe('<span style="font-size: 20px;">😊</span>')
        else:
            return mark_safe('<span style="font-size: 20px;">😡</span>')

    teams_status.short_description = "Teams"

    def moodle_status(self, obj):
        """Muestra estado de Moodle con emoticono."""
        from django.utils.safestring import mark_safe
        if obj.moodle_procesado:
            return mark_safe('<span style="font-size: 20px;">😊</span>')
        else:
            return mark_safe('<span style="font-size: 20px;">😡</span>')

    moodle_status.short_description = "Moodle"

    def email_status(self, obj):
        """Muestra estado de Email con emoticono."""
        from django.utils.safestring import mark_safe
        if obj.email_procesado:
            return mark_safe('<span style="font-size: 20px;">😊</span>')
        else:
            return mark_safe('<span style="font-size: 20px;">😡</span>')

    email_status.short_description = "Email"

    @admin.action(description="📥 Exportar alumnos seleccionados a Excel")
    def exportar_alumnos_excel(self, request, queryset):
        """
        Exporta los alumnos seleccionados o filtrados a un archivo Excel.
        """
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Alumnos"

        # Encabezados
        headers = [
            'Apellido', 'Nombre', 'Tipo Documento', 'DNI', 'Email Personal',
            'Email Institucional', 'Estado Actual', 'Fecha Ingreso', 'Cohorte',
            'Localidad', 'Teléfono', 'Modalidad', 'Carreras', 'Teams Procesado',
            'Moodle Procesado', 'Email Procesado', 'Fecha Creación', 'Última Modificación'
        ]

        # Estilo de encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Escribir datos
        for row_num, alumno in enumerate(queryset, 2):
            # Formatear carreras
            carreras_texto = ""
            if alumno.carreras_data:
                carreras_nombres = [c.get('nombre_carrera', 'N/A') for c in alumno.carreras_data]
                carreras_texto = ", ".join(carreras_nombres)

            # Fecha de ingreso
            fecha_ingreso = alumno.fecha_ingreso.strftime('%Y-%m-%d') if alumno.fecha_ingreso else ''

            # Fecha de creación y modificación
            fecha_creacion = alumno.created_at.strftime('%Y-%m-%d %H:%M') if alumno.created_at else ''
            fecha_modificacion = alumno.updated_at.strftime('%Y-%m-%d %H:%M') if alumno.updated_at else ''

            row_data = [
                alumno.apellido,
                alumno.nombre,
                alumno.tipo_documento,
                alumno.dni,
                alumno.email_personal,
                alumno.email_institucional or '',
                alumno.estado_actual,
                fecha_ingreso,
                alumno.cohorte or '',
                alumno.localidad or '',
                alumno.telefono or '',
                alumno.modalidad_actual or '',
                carreras_texto,
                'Sí' if alumno.teams_procesado else 'No',
                'Sí' if alumno.moodle_procesado else 'No',
                'Sí' if alumno.email_procesado else 'No',
                fecha_creacion,
                fecha_modificacion
            ]

            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"alumnos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Guardar workbook en respuesta
        wb.save(response)

        self.message_user(
            request,
            f"✅ Exportados {queryset.count()} alumnos a Excel",
            level=messages.SUCCESS
        )

        return response

    exportar_alumnos_excel.short_description = "📥 Exportar alumnos a Excel"

    def carreras_formatted(self, obj):
        """Muestra la carrera del alumno (primera carrera de la lista)."""
        if not obj.carreras_data or len(obj.carreras_data) == 0:
            return "No tiene carrera asignada"

        from django.utils.safestring import mark_safe
        from cursos.utils import get_carrera_nombre, get_carrera_codigo
        import json

        # Solo mostrar la primera carrera (caso normal: una sola carrera)
        carrera = obj.carreras_data[0]
        id_carrera = carrera.get("id_carrera", "N/A")

        # Obtener nombre real desde la base de datos
        if id_carrera != "N/A":
            nombre_carrera = get_carrera_nombre(id_carrera)
            codigo_carrera = get_carrera_codigo(id_carrera) or "N/A"
        else:
            nombre_carrera = carrera.get("nombre_carrera", "Sin nombre")
            codigo_carrera = "N/A"

        modalidad = carrera.get("modalidad", "N/A")
        modalidad_texto = "Presencial" if modalidad == "1" else "Distancia" if modalidad == "2" else modalidad
        fecha_inscri = carrera.get("fecha_inscri", "N/A")
        comisiones = carrera.get("comisiones", [])

        comisiones_texto = ", ".join([
            f"{c.get('nombre_comision', '')} (ID: {c.get('id_comision', '')})"
            for c in comisiones
        ]) if comisiones else "Sin comisiones"

        # Nota si hay más de una carrera (caso raro)
        nota_multiple = ""
        if len(obj.carreras_data) > 1:
            nota_multiple = f'<p style="color: #856404; background-color: #fff3cd; padding: 10px; border-radius: 5px; margin-top: 10px;"><strong>⚠️ Nota:</strong> Este alumno tiene {len(obj.carreras_data)} carreras. Solo se muestra la primera. Casos múltiples se gestionan manualmente.</p>'

        json_str = json.dumps(obj.carreras_data, indent=2, ensure_ascii=False)
        html = f"""
            <div style="border: 1px solid #ddd; padding: 15px; background-color: #f9f9f9; border-radius: 5px;">
                <p><strong>ID Carrera (UTI):</strong> {id_carrera}</p>
                <p><strong>Código:</strong> {codigo_carrera}</p>
                <p><strong>Nombre:</strong> {nombre_carrera}</p>
                <p><strong>Modalidad:</strong> {modalidad_texto} (código: {modalidad})</p>
                <p><strong>Fecha Inscripción:</strong> {fecha_inscri}</p>
                <p><strong>Comisiones:</strong> {comisiones_texto}</p>
            </div>
            {nota_multiple}
            <details style="margin-top: 10px;">
                <summary style="cursor: pointer; color: #417690;">Ver datos completos (JSON)</summary>
                <pre style="background-color: #f5f5f5; padding: 10px; border-radius: 5px; overflow-x: auto; font-size: 12px;">{json_str}</pre>
            </details>
        """

        return mark_safe(html)

    carreras_formatted.short_description = "Carrera (desde API UTI)"


class CarreraListFilter(admin.SimpleListFilter):
    """Filtro personalizado para filtrar por nombre de carrera."""
    title = 'Carrera'
    parameter_name = 'carrera'

    def lookups(self, request, model_admin):
        """Obtiene lista única de carreras de todos los alumnos."""
        from django.db.models import Q

        # Obtener todas las carreras únicas
        alumnos = Alumno.objects.exclude(carreras_data__isnull=True)
        carreras_set = set()

        for alumno in alumnos:
            if alumno.carreras_data:
                for carrera in alumno.carreras_data:
                    nombre = carrera.get("nombre_carrera")
                    id_carrera = carrera.get("id_carrera")
                    if nombre:
                        carreras_set.add((str(id_carrera), nombre))

        return sorted(carreras_set, key=lambda x: x[1])

    def queryset(self, request, queryset):
        """Filtra los alumnos que tienen la carrera seleccionada."""
        if self.value():
            # Buscar alumnos donde carreras_data contenga el id_carrera
            return queryset.filter(carreras_data__contains=[{"id_carrera": int(self.value())}])
        return queryset


# Re-registrar AlumnoAdmin con el filtro personalizado
admin.site.unregister(Alumno)

@admin.register(Alumno)
class AlumnoAdminWithFilters(AlumnoAdmin):
    """AlumnoAdmin con filtros personalizados."""
    list_filter = ("estado_actual", "modalidad_actual", CarreraListFilter, "cohorte")


@admin.register(Log)
class LogAdmin(admin.ModelAdmin):
    """Admin para visualizar logs del sistema."""

    list_display = (
        'fecha',
        'tipo_colored',
        'modulo',
        'mensaje_corto',
        'alumno_link',
        'usuario',
    )
    list_filter = (
        'tipo',
        'modulo',
        ('fecha', admin.DateFieldListFilter),
    )
    search_fields = (
        'mensaje',
        'modulo',
        'usuario',
        'alumno__nombre',
        'alumno__apellido',
        'alumno__dni',
    )
    readonly_fields = (
        'tipo',
        'modulo',
        'mensaje',
        'detalles_formatted',
        'alumno',
        'usuario',
        'fecha',
    )
    date_hierarchy = 'fecha'
    ordering = ('-fecha',)

    def has_add_permission(self, request):
        """No permitir agregar logs manualmente."""
        return False

    def has_change_permission(self, request, obj=None):
        """No permitir editar logs."""
        return False

    def has_delete_permission(self, request, obj=None):
        """Permitir eliminar logs (limpieza)."""
        return request.user.is_superuser

    def tipo_colored(self, obj):
        """Muestra el tipo con color."""
        colors = {
            'SUCCESS': 'green',
            'INFO': 'blue',
            'WARNING': 'orange',
            'ERROR': 'red',
            'DEBUG': 'gray',
        }
        color = colors.get(obj.tipo, 'black')
        return f'<span style="color: {color}; font-weight: bold;">{obj.get_tipo_display()}</span>'
    tipo_colored.short_description = 'Tipo'
    tipo_colored.allow_tags = True

    def mensaje_corto(self, obj):
        """Muestra los primeros 100 caracteres del mensaje."""
        if len(obj.mensaje) > 100:
            return obj.mensaje[:100] + '...'
        return obj.mensaje
    mensaje_corto.short_description = 'Mensaje'

    def alumno_link(self, obj):
        """Link al alumno si existe."""
        if obj.alumno:
            from django.urls import reverse
            from django.utils.html import format_html
            url = reverse('admin:alumnos_alumno_change', args=[obj.alumno.id])
            return format_html('<a href="{}">{}</a>', url, obj.alumno)
        return '-'
    alumno_link.short_description = 'Alumno'

    def detalles_formatted(self, obj):
        """Formatea el JSON de detalles para visualización."""
        if obj.detalles:
            import json
            from django.utils.html import format_html
            formatted = json.dumps(obj.detalles, indent=2, ensure_ascii=False)
            return format_html('<pre>{}</pre>', formatted)
        return '-'
    detalles_formatted.short_description = 'Detalles'


@admin.register(Tarea)
class TareaAdmin(admin.ModelAdmin):
    """Admin para visualizar tareas asíncronas de Celery."""

    list_display = (
        'id',
        'tipo_colored',
        'estado_colored',
        'hora_programada',
        'duracion_formatted',
        'cantidad_entidades',
        'alumno_link',
        'usuario',
    )
    list_filter = (
        'tipo',
        'estado',
        ('hora_programada', admin.DateFieldListFilter),
    )
    search_fields = (
        'celery_task_id',
        'alumno__nombre',
        'alumno__apellido',
        'alumno__dni',
        'usuario',
        'mensaje_error',
    )
    readonly_fields = (
        'tipo',
        'estado',
        'celery_task_id',
        'hora_programada',
        'hora_inicio',
        'hora_fin',
        'cantidad_entidades',
        'detalles_formatted',
        'alumno',
        'usuario',
        'mensaje_error',
        'duracion_formatted',
    )
    fieldsets = (
        ('Información de la Tarea', {
            'fields': ('tipo', 'estado', 'celery_task_id')
        }),
        ('Tiempos', {
            'fields': ('hora_programada', 'hora_inicio', 'hora_fin', 'duracion_formatted')
        }),
        ('Resultados', {
            'fields': ('cantidad_entidades', 'detalles_formatted', 'mensaje_error')
        }),
        ('Contexto', {
            'fields': ('alumno', 'usuario')
        }),
    )
    date_hierarchy = 'hora_programada'
    ordering = ('-hora_programada',)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def tipo_colored(self, obj):
        """Muestra el tipo de tarea con color."""
        from django.utils.html import format_html
        color_map = {
            'ingesta_preinscriptos': '#3498db',
            'ingesta_aspirantes': '#2ecc71',
            'ingesta_ingresantes': '#9b59b6',
            'eliminar_cuenta': '#e74c3c',
            'enviar_email': '#f39c12',
            'activar_servicios': '#1abc9c',
            'crear_usuario_teams': '#34495e',
            'resetear_password': '#e67e22',
        }
        color = color_map.get(obj.tipo, '#95a5a6')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_tipo_display()
        )
    tipo_colored.short_description = 'Tipo'
    tipo_colored.admin_order_field = 'tipo'

    def estado_colored(self, obj):
        """Muestra el estado con color."""
        from django.utils.html import format_html
        color_map = {
            'pending': '#f39c12',
            'running': '#3498db',
            'completed': '#27ae60',
            'failed': '#e74c3c',
        }
        icon_map = {
            'pending': '⏳',
            'running': '▶️',
            'completed': '✅',
            'failed': '❌',
        }
        color = color_map.get(obj.estado, '#95a5a6')
        icon = icon_map.get(obj.estado, '❓')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} {}</span>',
            color, icon, obj.get_estado_display()
        )
    estado_colored.short_description = 'Estado'
    estado_colored.admin_order_field = 'estado'

    def alumno_link(self, obj):
        """Link al alumno relacionado."""
        from django.urls import reverse
        from django.utils.html import format_html
        if obj.alumno:
            url = reverse('admin:alumnos_alumno_change', args=[obj.alumno.id])
            return format_html('<a href="{}">{}</a>', url, obj.alumno)
        return '-'
    alumno_link.short_description = 'Alumno'

    def detalles_formatted(self, obj):
        """Muestra detalles en formato JSON."""
        if obj.detalles:
            import json
            from django.utils.html import format_html
            formatted = json.dumps(obj.detalles, indent=2, ensure_ascii=False)
            return format_html('<pre>{}</pre>', formatted)
        return '-'
    detalles_formatted.short_description = 'Detalles'

    def duracion_formatted(self, obj):
        """Muestra la duración de la tarea."""
        from django.utils.html import format_html
        duracion = obj.duracion
        if duracion is not None:
            # format_html no soporta format specs, formatear primero
            duracion_str = f'{duracion:.2f}'
            return format_html('{} seg', duracion_str)
        elif obj.estado == 'running':
            return format_html('<span style="color: #3498db;">En progreso...</span>')
        return '-'
    duracion_formatted.short_description = 'Duración'


@admin.register(Configuracion)
class ConfiguracionAdmin(admin.ModelAdmin):
    """Admin para configuración del sistema (Singleton)."""

    fieldsets = (
        ('⚙️ Procesamiento en Lotes y Rate Limiting', {
            'fields': (
                'batch_size',
                'rate_limit_teams',
                'rate_limit_moodle',
                'rate_limit_uti',
            ),
            'description': '🔧 Configuración de workflows automáticos. BATCH_SIZE: alumnos por lote. RATE_LIMIT: tareas/llamadas máximas por minuto para cada servicio (Teams, Moodle, UTI).'
        }),
        ('📥 Ingesta Automática - Preinscriptos', {
            'fields': (
                'preinscriptos_dia_inicio',
                'preinscriptos_dia_fin',
                'preinscriptos_frecuencia_segundos',
                'preinscriptos_forzar_carga_completa',
                'preinscriptos_enviar_email',
                'preinscriptos_activar_teams',
                'preinscriptos_activar_moodle',
            ),
            'description': '✉️ Configuración de ingesta automática de preinscriptos. FORZAR CARGA COMPLETA: trae todos los registros desde dia_inicio (se desactiva automáticamente después de ejecutar). Los checkboxes controlan emails y activación automática en Teams/Moodle.'
        }),
        ('📥 Ingesta Automática - Aspirantes', {
            'fields': (
                'aspirantes_dia_inicio',
                'aspirantes_dia_fin',
                'aspirantes_frecuencia_segundos',
                'aspirantes_forzar_carga_completa',
                'aspirantes_enviar_email',
                'aspirantes_activar_teams',
                'aspirantes_activar_moodle',
            ),
            'description': '✉️ Configuración de ingesta automática de aspirantes. FORZAR CARGA COMPLETA: trae todos los registros desde dia_inicio (se desactiva automáticamente después de ejecutar). Los checkboxes controlan emails y activación automática en Teams/Moodle.'
        }),
        ('📥 Ingesta Automática - Ingresantes', {
            'fields': (
                'ingresantes_dia_inicio',
                'ingresantes_dia_fin',
                'ingresantes_frecuencia_segundos',
                'ingresantes_forzar_carga_completa',
                'ingresantes_enviar_email',
                'ingresantes_activar_teams',
                'ingresantes_activar_moodle',
            ),
            'description': '✉️ Configuración de ingesta automática de ingresantes. FORZAR CARGA COMPLETA: trae todos los registros desde dia_inicio (se desactiva automáticamente después de ejecutar). Los checkboxes controlan emails y activación automática en Teams/Moodle.'
        }),
        ('🔐 Credenciales Teams/Azure AD', {
            'fields': (
                'teams_tenant_id',
                'teams_client_id',
                'teams_client_secret',
                'account_prefix',
            ),
            'description': 'Credenciales de Teams y prefijo de cuentas. Si están vacías, se usan las variables de entorno. ACCOUNT_PREFIX: "test-a" para testing, "a" para producción.',
            'classes': ('collapse',)
        }),
        ('🏛️ API SIAL/UTI', {
            'fields': (
                'sial_base_url',
                'sial_basic_user',
                'sial_basic_pass',
            ),
            'description': 'Configuración de API SIAL/UTI. URL: https://sial.unrc.edu.ar (producción) o http://mock-api-uti:8000 (mock). Credenciales para autenticación básica. Si están vacíos, usan variables de entorno.',
            'classes': ('collapse',)
        }),
        ('🎓 Credenciales Moodle', {
            'fields': (
                'moodle_base_url',
                'moodle_wstoken',
                'moodle_email_type',
                'moodle_student_roleid',
                'moodle_auth_method',
                'moodle_courses_config',
            ),
            'description': 'Credenciales de Moodle. Auth method: oauth2 (Microsoft) o manual. Courses config: JSON con cursos por estado de alumno.',
            'classes': ('collapse',)
        }),
        ('📧 Configuración de Email', {
            'fields': (
                'email_usar_microsoft_graph',
                'email_from',
                'email_host',
                'email_port',
                'email_use_tls',
                'deshabilitar_fallback_email_personal',
            ),
            'description': 'Configuración de envío de emails. MICROSOFT GRAPH: Si se activa, usa Microsoft Graph API en lugar de SMTP (requiere credenciales Teams configuradas, solo necesita email_from). SMTP: Si Graph está desactivado, usa configuración SMTP tradicional. ⚠️ FALLBACK: Si está deshabilitado, el sistema NO usará email_personal cuando falte email_institucional.',
            'classes': ('collapse',)
        }),
        ('✉️ Plantillas de Emails', {
            'fields': (
                'email_asunto_bienvenida',
                'email_plantilla_bienvenida',
                'email_asunto_credenciales',
                'email_plantilla_credenciales',
                'email_asunto_password',
                'email_plantilla_password',
                'email_asunto_enrollamiento',
                'email_plantilla_enrollamiento',
            ),
            'description': 'Plantillas HTML para emails. Pegar HTML completo en cada campo. Variables disponibles: {nombre}, {apellido}, {dni}, {email}, {upn}, {password}, {moodle_url}, {cursos_html}. IMPORTANTE: En CSS usar {{{{ y }}}} para escapar llaves.',
            'classes': ('collapse',)
        }),
        ('ℹ️ Metadatos', {
            'fields': (
                'actualizado_en',
                'actualizado_por',
            ),
            'classes': ('collapse',)
        }),
    )

    readonly_fields = ('actualizado_en', 'actualizado_por')

    actions = ['resetear_checkpoints_ingesta']

    def has_add_permission(self, request):
        """Solo puede haber una configuración (Singleton)."""
        return not Configuracion.objects.exists()

    def has_delete_permission(self, request, obj=None):
        """No se puede eliminar la configuración."""
        return False

    def save_model(self, request, obj, form, change):
        """Guarda quién modificó la configuración."""
        obj.actualizado_por = request.user.username
        super().save_model(request, obj, form, change)

    @admin.action(description="🔄 Resetear checkpoints de ingesta (fuerza carga completa en próxima ejecución)")
    def resetear_checkpoints_ingesta(self, request, queryset):
        """Resetea los timestamps de última ingesta para forzar carga completa."""
        for config in queryset:
            config.ultima_ingesta_preinscriptos = None
            config.ultima_ingesta_aspirantes = None
            config.ultima_ingesta_ingresantes = None
            config.save(update_fields=[
                'ultima_ingesta_preinscriptos',
                'ultima_ingesta_aspirantes',
                'ultima_ingesta_ingresantes'
            ])

        self.message_user(
            request,
            "Checkpoints reseteados. La próxima ingesta automática traerá todos los registros desde dia_inicio.",
            messages.SUCCESS
        )

    def get_urls(self):
        """Agregar URLs personalizadas para exportar/importar."""
        urls = super().get_urls()
        from django.urls import path
        custom_urls = [
            path(
                'exportar-json/',
                self.admin_site.admin_view(self.exportar_json_view),
                name='alumnos_configuracion_exportar_json',
            ),
            path(
                'importar-json/',
                self.admin_site.admin_view(self.importar_json_view),
                name='alumnos_configuracion_importar_json',
            ),
        ]
        return custom_urls + urls

    def exportar_json_view(self, request):
        """Exporta la configuración a JSON."""
        from django.http import HttpResponse
        from django.core.serializers.json import DjangoJSONEncoder
        import json
        from datetime import datetime, time

        config = Configuracion.load()

        # Función helper para convertir time a string
        def time_to_str(t):
            if isinstance(t, time):
                return t.strftime('%H:%M:%S')
            return t

        # Crear diccionario con todos los campos excepto id, actualizado_en, actualizado_por
        data = {
            'batch_size': config.batch_size,
            'rate_limit_teams': config.rate_limit_teams,
            'rate_limit_moodle': config.rate_limit_moodle,
            'rate_limit_uti': config.rate_limit_uti,
            'preinscriptos_dia_inicio': time_to_str(config.preinscriptos_dia_inicio),
            'preinscriptos_dia_fin': time_to_str(config.preinscriptos_dia_fin),
            'preinscriptos_frecuencia_segundos': config.preinscriptos_frecuencia_segundos,
            'preinscriptos_enviar_email': config.preinscriptos_enviar_email,
            'preinscriptos_activar_teams': config.preinscriptos_activar_teams,
            'preinscriptos_activar_moodle': config.preinscriptos_activar_moodle,
            'aspirantes_dia_inicio': time_to_str(config.aspirantes_dia_inicio),
            'aspirantes_dia_fin': time_to_str(config.aspirantes_dia_fin),
            'aspirantes_frecuencia_segundos': config.aspirantes_frecuencia_segundos,
            'aspirantes_enviar_email': config.aspirantes_enviar_email,
            'aspirantes_activar_teams': config.aspirantes_activar_teams,
            'aspirantes_activar_moodle': config.aspirantes_activar_moodle,
            'ingresantes_dia_inicio': time_to_str(config.ingresantes_dia_inicio),
            'ingresantes_dia_fin': time_to_str(config.ingresantes_dia_fin),
            'ingresantes_frecuencia_segundos': config.ingresantes_frecuencia_segundos,
            'ingresantes_enviar_email': config.ingresantes_enviar_email,
            'ingresantes_activar_teams': config.ingresantes_activar_teams,
            'ingresantes_activar_moodle': config.ingresantes_activar_moodle,
            'teams_tenant_id': config.teams_tenant_id,
            'teams_client_id': config.teams_client_id,
            'teams_client_secret': config.teams_client_secret,
            'account_prefix': config.account_prefix,
            'sial_base_url': config.sial_base_url,
            'sial_basic_user': config.sial_basic_user,
            'sial_basic_pass': config.sial_basic_pass,
            'moodle_base_url': config.moodle_base_url,
            'moodle_wstoken': config.moodle_wstoken,
            'moodle_email_type': config.moodle_email_type,
            'moodle_student_roleid': config.moodle_student_roleid,
            'moodle_auth_method': config.moodle_auth_method,
            'moodle_courses_config': config.moodle_courses_config,
            'email_from': config.email_from,
            'email_host': config.email_host,
            'email_port': config.email_port,
            'email_use_tls': config.email_use_tls,
            'email_asunto_bienvenida': config.email_asunto_bienvenida,
            'email_plantilla_bienvenida': config.email_plantilla_bienvenida,
            'email_asunto_credenciales': config.email_asunto_credenciales,
            'email_plantilla_credenciales': config.email_plantilla_credenciales,
            'email_asunto_password': config.email_asunto_password,
            'email_plantilla_password': config.email_plantilla_password,
            'email_asunto_enrollamiento': config.email_asunto_enrollamiento,
            'email_plantilla_enrollamiento': config.email_plantilla_enrollamiento,
        }

        # Crear respuesta JSON usando DjangoJSONEncoder para manejar datetime/time
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False, cls=DjangoJSONEncoder),
            content_type='application/json'
        )
        filename = f"configuracion_pylucy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        messages.success(request, "✅ Configuración exportada correctamente")
        return response

    def importar_json_view(self, request):
        """Importa la configuración desde JSON."""
        from django.shortcuts import redirect
        from django.urls import reverse
        from datetime import datetime
        import json

        if request.method != 'POST':
            messages.error(request, "❌ Método no permitido")
            return redirect('..')

        if 'json_file' not in request.FILES:
            messages.error(request, "❌ No se seleccionó ningún archivo")
            return redirect('..')

        try:
            json_file = request.FILES['json_file']
            data = json.load(json_file)

            config = Configuracion.load()

            # Campos que son de tipo time y necesitan conversión
            time_fields = [
                'preinscriptos_dia_inicio', 'preinscriptos_dia_fin',
                'aspirantes_dia_inicio', 'aspirantes_dia_fin',
                'ingresantes_dia_inicio', 'ingresantes_dia_fin'
            ]

            # Actualizar todos los campos
            for field, value in data.items():
                if hasattr(config, field):
                    # Convertir strings de tiempo a objetos time
                    if field in time_fields and isinstance(value, str):
                        value = datetime.strptime(value, '%H:%M:%S').time()
                    setattr(config, field, value)

            config.actualizado_por = request.user.username
            config.save()

            messages.success(
                request,
                f"✅ Configuración importada correctamente desde {json_file.name}"
            )

        except json.JSONDecodeError as e:
            messages.error(request, f"❌ Error al parsear JSON: {str(e)}")
        except Exception as e:
            messages.error(request, f"❌ Error al importar: {str(e)}")

        return redirect('..')


    def changelist_view(self, request, extra_context=None):
        """Redirige directamente al formulario de edición."""
        obj = Configuracion.load()
        from django.urls import reverse
        from django.http import HttpResponseRedirect
        url = reverse('admin:alumnos_configuracion_change', args=[obj.pk])
        return HttpResponseRedirect(url)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Agrega botones de exportar/importar al formulario."""
        extra_context = extra_context or {}
        extra_context['show_export_import'] = True
        return super().change_view(request, object_id, form_url, extra_context)

# ============================================
# VISTA PERSONALIZADA DEL ADMIN INDEX
# ============================================

from django.contrib.admin import AdminSite
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail
from django.conf import settings


class PyLucyAdminSite(AdminSite):
    """Admin personalizado con dashboard extendido."""
    site_header = "PyLucy - Sistema de Gestión de Alumnos"
    site_title = "PyLucy Admin"
    index_title = "Dashboard"
    index_template = "admin/pylucy_index.html"
    enable_nav_sidebar = True  # Habilitar sidebar de navegación

    def index(self, request, extra_context=None):
        """Vista del dashboard personalizado con datos de tareas."""
        from django.utils import timezone
        from datetime import timedelta

        # Obtener tareas recientes
        tareas_recientes = Tarea.objects.all().order_by('-hora_programada')[:10]

        # Obtener resumen de tareas
        now = timezone.now()
        last_24h = now - timedelta(hours=24)

        resumen = {
            'pending': Tarea.objects.filter(estado='pending').count(),
            'running': Tarea.objects.filter(estado='running').count(),
            'completed': Tarea.objects.filter(estado='completed', hora_programada__gte=last_24h).count(),
            'failed': Tarea.objects.filter(estado='failed', hora_programada__gte=last_24h).count(),
        }

        context = {
            'tareas_recientes': tareas_recientes,
            'resumen': resumen,
            **(extra_context or {}),
        }

        return super().index(request, extra_context=context)

    def get_urls(self):
        """Agregar URLs personalizadas."""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('probar-email/', self.admin_view(self.test_email_view), name='test_email'),
        ]
        return custom_urls + urls

    def test_email_view(self, request):
        """Vista para probar envío de emails usando Microsoft Graph API."""
        # Si es GET, mostrar el formulario
        if request.method == 'GET':
            from django.shortcuts import render
            return render(request, 'admin/test_email.html')

        # Si es POST, procesar el envío del email
        if request.method == 'POST':
            try:
                # Obtener parámetros del formulario
                destinatario = request.POST.get('destinatario', '').strip()
                email_from = request.POST.get('email_from', '').strip()
                mensaje_personalizado = request.POST.get('mensaje', '').strip()

                # Validaciones
                if not destinatario or not email_from:
                    return JsonResponse({
                        'success': False,
                        'error': 'Debe ingresar remitente y destinatario'
                    })

                # Importar TeamsService para usar el token de Graph API
                from .services.teams_service import TeamsService
                teams_svc = TeamsService()

                # Email de prueba
                subject = "✅ Prueba de Email Office 365 - PyLucy"

                # Mensaje personalizado o por defecto
                if mensaje_personalizado:
                    message_body = mensaje_personalizado
                else:
                    message_body = """Hola,

Este es un email de prueba enviado desde PyLucy usando Microsoft Graph API.

Si recibes este mensaje, la configuración de Office 365 está funcionando correctamente.

Saludos,
Sistema PyLucy"""

                html_message = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background-color: #0078d4; color: white; padding: 20px; text-align: center; }}
        .content {{ background-color: #f9f9f9; padding: 20px; }}
        .message-box {{ background-color: white; border: 1px solid #ddd; padding: 15px; margin: 20px 0; border-radius: 4px; white-space: pre-wrap; }}
        .info {{ background-color: #e8f4f8; border-left: 4px solid #0066cc; padding: 15px; margin: 20px 0; }}
        .footer {{ background-color: #f0f0f0; padding: 15px; text-align: center; font-size: 12px; color: #666; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>✅ Prueba de Email Office 365</h1>
            <p style="margin: 0;">PyLucy - Sistema de Gestión de Alumnos</p>
        </div>

        <div class="content">
            <div class="message-box">{message_body}</div>

            <div class="info">
                <h3 style="margin-top: 0;">📋 Detalles técnicos de envío</h3>
                <ul style="margin-bottom: 0;">
                    <li><strong>Método:</strong> Microsoft Graph API</li>
                    <li><strong>From:</strong> {email_from}</li>
                    <li><strong>To:</strong> {destinatario}</li>
                </ul>
            </div>

            <p>Si recibes este mensaje, Microsoft Graph API está configurado correctamente.</p>

            <p>Saludos,<br>
            <strong>Sistema PyLucy</strong></p>
        </div>

        <div class="footer">
            <p>Este es un mensaje de prueba del sistema PyLucy</p>
        </div>
    </div>
</body>
</html>
"""

                # Enviar usando Microsoft Graph API
                import requests

                # Obtener token
                token = teams_svc._get_token()

                # Construir mensaje
                message_payload = {
                    "message": {
                        "subject": subject,
                        "body": {
                            "contentType": "HTML",
                            "content": html_message
                        },
                        "toRecipients": [
                            {
                                "emailAddress": {
                                    "address": destinatario
                                }
                            }
                        ]
                    },
                    "saveToSentItems": "true"
                }

                # Enviar email usando Graph API
                # https://graph.microsoft.com/v1.0/users/{userId}/sendMail
                url = f"https://graph.microsoft.com/v1.0/users/{email_from}/sendMail"

                logger.info(f"Enviando email vía Graph API desde {email_from} a {destinatario}")

                response = requests.post(
                    url,
                    json=message_payload,
                    headers={
                        'Authorization': f'Bearer {token}',
                        'Content-Type': 'application/json'
                    },
                    timeout=30
                )

                response.raise_for_status()

                logger.info(f"✅ Email de prueba enviado exitosamente a {destinatario}")
                return JsonResponse({
                    'success': True,
                    'message': f'Email enviado exitosamente a {destinatario} usando Microsoft Graph API'
                })

            except requests.exceptions.HTTPError as e:
                error_msg = f"Error HTTP {e.response.status_code}: {e.response.text}"
                logger.error(error_msg)

                hint = None
                if e.response.status_code == 403:
                    hint = 'Verifica que el permiso Mail.Send esté configurado y tenga admin consent.'
                elif e.response.status_code == 404:
                    hint = f'El buzón "{email_from}" no existe en Office 365.'

                return JsonResponse({
                    'success': False,
                    'error': error_msg,
                    'hint': hint
                })

            except Exception as e:
                logger.error(f"Error enviando email de prueba: {e}")
                import traceback
                return JsonResponse({
                    'success': False,
                    'error': str(e),
                    'hint': 'Verifica la configuración de Azure AD (TEAMS_TENANT, TEAMS_CLIENT_ID, TEAMS_CLIENT_SECRET)'
                })

        return JsonResponse({'success': False, 'error': 'Método no permitido'})


# Reemplazar el admin site por defecto
admin_site = PyLucyAdminSite(name='admin')

# Re-registrar todos los modelos en el nuevo admin site
admin_site.register(Alumno, AlumnoAdmin)
admin_site.register(Log, LogAdmin)
admin_site.register(Configuracion, ConfiguracionAdmin)
admin_site.register(Tarea, TareaAdmin)

# Registrar modelos de la app cursos
from cursos.models import CursoIngreso, Carrera
from cursos.admin import CursoIngresoAdmin, CarreraAdmin
admin_site.register(CursoIngreso, CursoIngresoAdmin)
admin_site.register(Carrera, CarreraAdmin)


# =============================================================================
# TAREAS PERSONALIZADAS
# =============================================================================
from .models import TareaPersonalizada


@admin.register(TareaPersonalizada, site=admin_site)
class TareaPersonalizadaAdmin(admin.ModelAdmin):
    """Admin para gestionar tareas personalizadas con periodicidad configurable."""

    list_display = (
        'nombre',
        'activa',
        'tipo_usuario',
        'accion',
        'respetar_rate_limits',
        'cantidad_ejecuciones',
        'ultima_ejecucion',
        'ver_periodic_task'
    )
    list_filter = ('activa', 'tipo_usuario', 'accion', 'respetar_rate_limits')
    search_fields = ('nombre',)
    readonly_fields = ('cantidad_ejecuciones', 'ultima_ejecucion', 'creada_en', 'modificada_en')

    fieldsets = (
        ('Información General', {
            'fields': ('nombre', 'activa')
        }),
        ('Configuración de la Tarea', {
            'fields': ('tipo_usuario', 'accion', 'respetar_rate_limits')
        }),
        ('Filtros de Fecha (solo para ingesta SIAL)', {
            'fields': ('fecha_desde', 'fecha_hasta'),
            'description': 'Configura el rango de fechas para consultar datos desde SIAL/UTI. Deja "fecha_hasta" vacío para usar la fecha actual.'
        }),
        ('Opciones de Procesamiento', {
            'fields': ('enviar_email',),
            'description': 'Configuración adicional para el procesamiento de usuarios.'
        }),
        ('Periodicidad', {
            'fields': ('periodic_task',),
            'description': 'Vincula esta tarea con una tarea periódica de Celery Beat. Crea primero un Crontab Schedule y luego una Periodic Task que ejecute "alumnos.tasks.ejecutar_tarea_personalizada".'
        }),
        ('Estadísticas', {
            'fields': ('cantidad_ejecuciones', 'ultima_ejecucion', 'creada_en', 'modificada_en'),
            'classes': ('collapse',)
        }),
    )

    def ver_periodic_task(self, obj):
        """Link a la tarea periódica asociada."""
        if obj.periodic_task:
            url = f'/admin/django_celery_beat/periodictask/{obj.periodic_task.id}/change/'
            return format_html('<a href="{}">Ver configuración</a>', url)
        return '-'
    ver_periodic_task.short_description = 'Periodicidad'


# =============================================================================
# DJANGO-CELERY-BEAT: Tareas Periódicas Configurables desde Admin
# =============================================================================
# Registramos lo necesario para configurar tareas periódicas:
# - PeriodicTask: Para crear/editar tareas programadas
# - CrontabSchedule: Para definir intervalos tipo cron (ej: */5 * * * *)
# - IntervalSchedule: Para intervalos en segundos/minutos/horas (ej: cada 1 segundo)
# =============================================================================
try:
    from django_celery_beat.models import (
        PeriodicTask, CrontabSchedule, IntervalSchedule
    )
    from django_celery_beat.admin import (
        PeriodicTaskAdmin as BasePeriodicTaskAdmin,
        CrontabScheduleAdmin as BaseCrontabScheduleAdmin,
        IntervalScheduleAdmin as BaseIntervalScheduleAdmin
    )

    # Personalizar PeriodicTaskAdmin para asegurar que se puedan agregar tareas
    class PyLucyPeriodicTaskAdmin(BasePeriodicTaskAdmin):
        """Admin personalizado para PeriodicTask con permisos completos."""

        def has_add_permission(self, request):
            """Permitir agregar tareas periódicas."""
            return True

        def has_change_permission(self, request, obj=None):
            """Permitir editar tareas periódicas."""
            return True

        def has_delete_permission(self, request, obj=None):
            """Permitir eliminar tareas periódicas."""
            return True

        def has_module_permission(self, request):
            """Mostrar el módulo en el admin."""
            return True

    # Personalizar CrontabScheduleAdmin
    class PyLucyCrontabScheduleAdmin(BaseCrontabScheduleAdmin):
        """Admin personalizado para CrontabSchedule con permisos completos."""

        def has_add_permission(self, request):
            """Permitir agregar crontabs."""
            return True

        def has_change_permission(self, request, obj=None):
            """Permitir editar crontabs."""
            return True

        def has_delete_permission(self, request, obj=None):
            """Permitir eliminar crontabs."""
            return True

        def has_module_permission(self, request):
            """Mostrar el módulo en el admin."""
            return True

    # Personalizar IntervalScheduleAdmin
    class PyLucyIntervalScheduleAdmin(BaseIntervalScheduleAdmin):
        """Admin personalizado para IntervalSchedule con permisos completos."""

        def has_add_permission(self, request):
            """Permitir agregar intervalos."""
            return True

        def has_change_permission(self, request, obj=None):
            """Permitir editar intervalos."""
            return True

        def has_delete_permission(self, request, obj=None):
            """Permitir eliminar intervalos."""
            return True

        def has_module_permission(self, request):
            """Mostrar el módulo en el admin."""
            return True

    # Registrar con admins personalizados
    admin_site.register(PeriodicTask, PyLucyPeriodicTaskAdmin)
    admin_site.register(CrontabSchedule, PyLucyCrontabScheduleAdmin)
    admin_site.register(IntervalSchedule, PyLucyIntervalScheduleAdmin)

except ImportError:
    # django-celery-beat no está instalado
    pass
